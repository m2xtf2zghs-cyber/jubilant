from __future__ import annotations

from copy import copy
import datetime as dt
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence

from dateutil import parser as date_parser
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


INDIAN_NUMBER_FORMAT = "#,##,##0.00"
DATE_NUMBER_FORMAT = "DD-MMM-YYYY"


def _copy_row_style(ws, src_row: int, dst_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)


def _unique_sheet_title(workbook, desired_title: str) -> str:
    base = (desired_title or "SHEET").strip()[:31] or "SHEET"
    if base not in workbook.sheetnames:
        return base
    suffix = 1
    while True:
        suffix_text = f"-{suffix}"
        candidate = f"{base[:31 - len(suffix_text)]}{suffix_text}"
        if candidate not in workbook.sheetnames:
            return candidate
        suffix += 1


def _as_excel_date(value: Any):
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, str):
        try:
            return date_parser.parse(value, dayfirst=True).date()
        except Exception:
            return value
    return value


def _month_label_from_key(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    try:
        d = dt.datetime.strptime(text, "%Y-%m")
        return d.strftime("%b-%y").upper()
    except Exception:
        return text


def fill_xns_sheet(ws, start_row: int, txns: List[Dict[str, Any]], template_row: int) -> None:
    """
    Write XNS rows while preserving template styles.
    Extra columns are additive: finance tag + confidence + reason codes.
    """
    max_col = max(ws.max_column, 13)
    for i, txn in enumerate(txns):
        row = start_row + i
        if row > ws.max_row:
            ws.insert_rows(row)
        _copy_row_style(ws, template_row, row, ws.max_column)

        ws.cell(row, 1).value = i + 1
        date_cell = ws.cell(row, 2)
        date_cell.value = _as_excel_date(txn.get("date"))
        if isinstance(date_cell.value, dt.date):
            date_cell.number_format = DATE_NUMBER_FORMAT

        ws.cell(row, 3).value = txn.get("month_label", "")
        ws.cell(row, 4).value = txn.get("txn_type", "")
        ws.cell(row, 5).value = txn.get("ref_no", "")
        ws.cell(row, 6).value = txn.get("category", "")
        ws.cell(row, 7).value = txn.get("narration", "")

        for col, key in ((8, "dr"), (9, "cr"), (10, "balance")):
            cell = ws.cell(row, col)
            cell.value = float(txn.get(key) or 0)
            cell.number_format = INDIAN_NUMBER_FORMAT

        ws.cell(row, 11).value = txn.get("finance_tag") or ""
        conf_cell = ws.cell(row, 12)
        conf_cell.value = float(txn.get("tag_confidence") or 0)
        conf_cell.number_format = "0.00000"
        ws.cell(row, 13).value = txn.get("reason_codes") or ""


def fill_pivot_sheet(ws, start_row: int, pivot_rows: List[Dict[str, Any]], template_row: int) -> None:
    """
    Fill pivot sheet rows with a style-preserving copy pattern.
    Expected keys: month_key, category, txn_type, sum_dr, sum_cr, count_dr, count_cr.
    """
    max_col = ws.max_column
    for i, pivot in enumerate(pivot_rows):
        row = start_row + i
        if row > ws.max_row:
            ws.insert_rows(row)
        _copy_row_style(ws, template_row, row, max_col)

        ws.cell(row, 1).value = _month_label_from_key(pivot.get("month_key", ""))
        ws.cell(row, 2).value = pivot.get("category", "")
        ws.cell(row, 3).value = pivot.get("txn_type", "")

        for col, key in ((4, "sum_dr"), (5, "sum_cr")):
            c = ws.cell(row, col)
            c.value = float(pivot.get(key) or 0)
            c.number_format = INDIAN_NUMBER_FORMAT

        ws.cell(row, 6).value = int(pivot.get("count_dr") or 0)
        ws.cell(row, 7).value = int(pivot.get("count_cr") or 0)


def _fill_value_rows(ws, start_row: int, rows: Iterable[Iterable[Any]]) -> None:
    for offset, row_values in enumerate(rows):
        row_no = start_row + offset
        for col_no, value in enumerate(row_values, start=1):
            ws.cell(row_no, col_no).value = value


def _prepare_simple_sheet(wb, sheet_name: str, headers: Sequence[str]):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(title=sheet_name)

    ws.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="F1F5F9")
    header_font = Font(bold=True)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
    return ws


def _write_simple_rows(
    ws,
    rows: Sequence[Sequence[Any]],
    date_cols: Sequence[int] = (),
    number_cols: Sequence[int] = (),
) -> None:
    for idx, row in enumerate(rows, start=2):
        for col, value in enumerate(row, start=1):
            cell = ws.cell(idx, col)
            if col in date_cols:
                parsed = _as_excel_date(value)
                cell.value = parsed
                if isinstance(parsed, dt.date):
                    cell.number_format = DATE_NUMBER_FORMAT
            else:
                cell.value = value

            if col in number_cols and isinstance(value, (int, float)):
                cell.number_format = INDIAN_NUMBER_FORMAT


def _write_totals_row(ws, row_no: int, label_col: int, values: Sequence[float], start_col: int) -> None:
    ws.cell(row_no, label_col).value = "TOTAL"
    ws.cell(row_no, label_col).font = Font(bold=True)
    for i, v in enumerate(values):
        cell = ws.cell(row_no, start_col + i)
        cell.value = float(v)
        cell.number_format = INDIAN_NUMBER_FORMAT
        cell.font = Font(bold=True)


def generate_perfios_excel(template_path: str, output_path: str, context: Dict[str, Any]) -> None:
    """
    Clone formatted template sheets and write additive analysis outputs.
    Heatmap/pivot logic stays unchanged; this writer only exports already-computed rows.
    """
    template = Path(template_path)
    if not template.exists():
        raise FileNotFoundError(f"Template workbook not found: {template}")

    wb = load_workbook(template)

    for account in context.get("accounts", []):
        xns_tpl_name = account["xns_template_sheet"]
        if xns_tpl_name not in wb.sheetnames:
            raise ValueError(f"XNS template sheet not found: {xns_tpl_name}")

        xns_template_ws = wb[xns_tpl_name]
        xns_ws = wb.copy_worksheet(xns_template_ws)
        xns_ws.title = _unique_sheet_title(wb, account.get("xns_sheet_name") or f"XNS-{len(wb.sheetnames)}")
        fill_xns_sheet(
            xns_ws,
            start_row=int(account.get("xns_start_row", 10)),
            txns=account.get("txns", []),
            template_row=int(account.get("xns_template_row", 10)),
        )

        pivot_tpl_name = account.get("pivot_template_sheet")
        if pivot_tpl_name and pivot_tpl_name in wb.sheetnames:
            pivot_template_ws = wb[pivot_tpl_name]
            pivot_ws = wb.copy_worksheet(pivot_template_ws)
            pivot_ws.title = _unique_sheet_title(
                wb,
                account.get("pivot_sheet_name") or f"PIVOT-{len(wb.sheetnames)}",
            )
            fill_pivot_sheet(
                pivot_ws,
                start_row=int(account.get("pivot_start_row", 2)),
                pivot_rows=account.get("pivots", []),
                template_row=int(account.get("pivot_template_row", 2)),
            )

    analysis_rows = context.get("analysis_rows") or []
    if analysis_rows:
        analysis_ws = wb["ANALYSIS"] if "ANALYSIS" in wb.sheetnames else _prepare_simple_sheet(wb, "ANALYSIS", ["Metric", "Value"])
        if analysis_ws.max_row <= 1:
            _prepare_simple_sheet(wb, "ANALYSIS", ["Metric", "Value"])
            analysis_ws = wb["ANALYSIS"]
        _fill_value_rows(analysis_ws, int(context.get("analysis_start_row", 2)), analysis_rows)

    cons_rows = context.get("cons_rows") or []
    if cons_rows:
        ws = _prepare_simple_sheet(wb, "CONS", ["Month", "Total DR", "Total CR", "Net"])
        rows = [
            [
                _month_label_from_key(r.get("month_key", "")),
                float(r.get("total_dr") or 0),
                float(r.get("total_cr") or 0),
                float(r.get("net") or 0),
            ]
            for r in cons_rows
        ]
        _write_simple_rows(ws, rows, number_cols=(2, 3, 4))
        total_row = len(rows) + 2
        _write_totals_row(
            ws,
            row_no=total_row,
            label_col=1,
            values=[sum(r[1] for r in rows), sum(r[2] for r in rows), sum(r[3] for r in rows)],
            start_col=2,
        )

    pvt_rows = context.get("pvt_fin_rows") or []
    if pvt_rows:
        ws = _prepare_simple_sheet(wb, "PVT FIN", ["Date", "Month", "TYPE", "Category", "DR", "CR", "Narration"])
        rows = [
            [
                r.get("date"),
                r.get("month_label", ""),
                r.get("type", "PVT FIN"),
                r.get("category", ""),
                float(r.get("dr") or 0),
                float(r.get("cr") or 0),
                r.get("narration", ""),
            ]
            for r in pvt_rows
        ]
        _write_simple_rows(ws, rows, date_cols=(1,), number_cols=(5, 6))
        total_row = len(rows) + 2
        _write_totals_row(
            ws,
            row_no=total_row,
            label_col=4,
            values=[sum(r[4] for r in rows), sum(r[5] for r in rows)],
            start_col=5,
        )

    bank_rows = context.get("bank_fin_rows") or []
    if bank_rows:
        ws = _prepare_simple_sheet(wb, "BANK FIN", ["Date", "Month", "TYPE", "Category", "DR", "CR", "Narration"])
        rows = [
            [
                r.get("date"),
                r.get("month_label", ""),
                r.get("type", "BANK FIN"),
                r.get("category", ""),
                float(r.get("dr") or 0),
                float(r.get("cr") or 0),
                r.get("narration", ""),
            ]
            for r in bank_rows
        ]
        _write_simple_rows(ws, rows, date_cols=(1,), number_cols=(5, 6))
        total_row = len(rows) + 2
        _write_totals_row(
            ws,
            row_no=total_row,
            label_col=4,
            values=[sum(r[4] for r in rows), sum(r[5] for r in rows)],
            start_col=5,
        )

    final_rows = context.get("final_rows") or []
    if final_rows:
        ws = _prepare_simple_sheet(wb, "FINAL", ["Field", "Value"])
        _write_simple_rows(ws, [list(r) for r in final_rows])

    wb.save(output_path)
