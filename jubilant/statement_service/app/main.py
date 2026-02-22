from __future__ import annotations

import datetime as dt
import hashlib
import os
import re
import tempfile
import uuid
from collections import Counter, defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple

from dateutil import parser as date_parser
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook

from .config import settings
from .excel.generate import generate_perfios_excel
from .parser.extract import extract_raw_lines_pdfplumber, merge_multiline_transactions
from .parser.reconcile import reconcile_strict
from .supabase_client import sb


app = FastAPI(title="Statement Autopilot Service", version="2.0.0")

_origins_raw = os.environ.get("STATEMENT_SERVICE_CORS_ORIGINS", "*")
_origins = [o.strip() for o in _origins_raw.split(",") if o.strip()] or ["*"]
app.add_middleware(
    CORSMiddleware,
    allow_origins=_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

AMOUNT_RE = re.compile(r"-?\d[\d,]*(?:\.\d{1,2})?")
NON_ALNUM_RE = re.compile(r"[^A-Z0-9 ]+")
SPACE_RE = re.compile(r"\s+")

DEFAULT_PVT_KEYWORDS = {
    "INTEREST": Decimal("0.75"),
    "WEEKLY": Decimal("0.70"),
    "BIWEEKLY": Decimal("0.65"),
    "FORTNIGHT": Decimal("0.65"),
    "COLLECTION": Decimal("0.45"),
    "LOAN": Decimal("0.35"),
    "FINANCE": Decimal("0.35"),
    "HAND LOAN": Decimal("1.10"),
    "PRIVATE": Decimal("0.80"),
    "MEDIATOR": Decimal("0.30"),
    "BROKER": Decimal("0.25"),
    "AGENT": Decimal("0.25"),
    "VATTI": Decimal("0.80"),
    "KANDHU": Decimal("0.80"),
    "METTU": Decimal("0.70"),
}


def _env_flag(name: str, default: bool) -> bool:
    raw = os.environ.get(name)
    if raw is None:
        return default
    return raw.strip().lower() not in {"0", "false", "no", "off", ""}

DEFAULT_BANK_KEYWORDS = {
    "EMI": Decimal("0.95"),
    "ECS": Decimal("0.90"),
    "NACH": Decimal("0.90"),
    "ACH": Decimal("0.85"),
    "AUTO DEBIT": Decimal("0.75"),
    "DISBURS": Decimal("0.70"),
    "LOAN A/C": Decimal("0.75"),
    "LOAN AC": Decimal("0.75"),
    "NBFC": Decimal("0.80"),
    "BANK": Decimal("0.45"),
    "MORATORIUM": Decimal("0.65"),
}

DEFAULT_FALSE_POSITIVES = {
    "SALARY",
    "REFUND",
    "REVERSAL",
    "INTERNAL TRANSFER",
    "SELF TRANSFER",
    "CASH DEPOSIT",
}

DEFAULT_TAG_THRESHOLDS = {
    "pvt_min_score": Decimal("2.10"),
    "bank_min_score": Decimal("2.40"),
    "weekly_window_days": 30,
    "weekly_min_hits": 3,
    "same_day_split_min_hits": 2,
    "small_ticket_max": Decimal("100000"),
}


class FinanceTagConfig(Dict[str, Any]):
    pvt_keywords: Dict[str, Decimal]
    bank_keywords: Dict[str, Decimal]
    false_patterns: Set[str]
    pvt_entities: Set[str]
    bank_entities: Set[str]
    thresholds: Dict[str, Any]


def _safe_table_select(table: str, select: str = "*", **kwargs: Any) -> List[Dict[str, Any]]:
    """Best-effort select that returns [] if a table is not available in current schema."""
    try:
        query = sb.table(table).select(select)
        if "eq" in kwargs:
            for col, value in kwargs["eq"].items():
                query = query.eq(col, value)
        if "order" in kwargs:
            order_col, ascending = kwargs["order"]
            query = query.order(order_col, desc=not ascending)
        if "limit" in kwargs:
            query = query.limit(int(kwargs["limit"]))
        resp = query.execute()
        return resp.data or []
    except Exception:
        return []


def _batch_insert(table: str, rows: List[Dict[str, Any]], size: int = 500) -> None:
    if not rows:
        return
    for i in range(0, len(rows), size):
        sb.table(table).insert(rows[i : i + size]).execute()


def _batch_upsert(table: str, rows: List[Dict[str, Any]], on_conflict: str, size: int = 500) -> None:
    if not rows:
        return
    for i in range(0, len(rows), size):
        sb.table(table).upsert(rows[i : i + size], on_conflict=on_conflict).execute()


def _parse_decimal(value: Any) -> Decimal:
    if value is None:
        return Decimal("0")
    text = str(value).replace(",", "").replace("₹", "").strip()
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def _infer_amount_triplet(dr_text: Any, cr_text: Any, bal_text: Any, raw_text: str) -> Tuple[Decimal, Decimal, Optional[Decimal]]:
    dr = _parse_decimal(dr_text)
    cr = _parse_decimal(cr_text)
    bal = _parse_decimal(bal_text) if bal_text not in (None, "") else None
    if dr or cr or bal is not None:
        return dr, cr, bal

    nums = [_parse_decimal(n) for n in AMOUNT_RE.findall(raw_text or "")]
    nums = [n for n in nums if n is not None]
    if not nums:
        return Decimal("0"), Decimal("0"), None
    if len(nums) == 1:
        return Decimal("0"), Decimal("0"), nums[0]
    if len(nums) == 2:
        return nums[0], Decimal("0"), nums[1]
    return nums[-3], nums[-2], nums[-1]


def _parse_date(value: str) -> Optional[dt.date]:
    if not value:
        return None
    try:
        return date_parser.parse(value, dayfirst=True).date()
    except Exception:
        return None


def _month_key(value: dt.date) -> str:
    return value.strftime("%Y-%m")


def _month_label(value: dt.date) -> str:
    return value.strftime("%b-%y").upper()


def _to_date_label(value: dt.date) -> str:
    return value.strftime("%d-%b-%Y").upper()


def _classify_txn_legacy(narration: str, dr: Decimal, cr: Decimal) -> str:
    """Existing category logic remains intact; finance_tag is additive."""
    text = (narration or "").upper()
    if any(k in text for k in ["RETURN", "RTN", "BOUNCE"]):
        return "RETURN"
    if any(k in text for k in ["EMI", "LOAN", "INTEREST", "OD INTEREST", "BANK CHARGES"]):
        return "BANK FIN"
    if any(k in text for k in ["PVT", "PRIVATE", "HAND LOAN"]):
        return "PVT FIN"
    amount = max(abs(dr), abs(cr))
    if amount >= Decimal("1000000") and (amount % 1000) != 0:
        return "ODD FIG"
    if "DOUBT" in text:
        return "DOUBT"
    if dr > 0 and cr > 0:
        return "CONS"
    return "FINAL"


def _txn_type(dr: Decimal, cr: Decimal) -> str:
    if cr > 0 and dr <= 0:
        return "CREDIT"
    if dr > 0 and cr <= 0:
        return "DEBIT"
    if dr > 0 and cr > 0:
        return "MIXED"
    return "UNKNOWN"


def _hash_uid(parts: Iterable[Any]) -> str:
    source = "|".join(str(p) for p in parts)
    return hashlib.sha1(source.encode("utf-8")).hexdigest()


def _normalize_text(value: str) -> str:
    t = str(value or "").upper()
    t = NON_ALNUM_RE.sub(" ", t)
    t = SPACE_RE.sub(" ", t).strip()
    return t


def _normalize_counterparty(narration: str) -> str:
    text = _normalize_text(narration)
    if not text:
        return "UNKNOWN"
    tokens = [t for t in text.split(" ") if len(t) >= 3]
    if not tokens:
        return "UNKNOWN"
    return " ".join(tokens[:3])


def _build_monthly_aggregates(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[row["month_key"]].append(row)

    aggregates = []
    for month, txns in sorted(grouped.items()):
        credit_total = sum(Decimal(str(t["cr"])) for t in txns)
        debit_total = sum(Decimal(str(t["dr"])) for t in txns)
        aggregates.append(
            {
                "month_key": month,
                "kpis": {
                    "txn_count": len(txns),
                    "credit_total": float(credit_total),
                    "debit_total": float(debit_total),
                    "net_flow": float(credit_total - debit_total),
                },
            }
        )
    return aggregates


def _build_pivot_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[Tuple[str, str, str], Dict[str, Any]] = {}
    for row in rows:
        key = (row["month_key"], row.get("category") or "", row.get("txn_type") or "")
        if key not in grouped:
            grouped[key] = {
                "month_key": key[0],
                "category": key[1],
                "txn_type": key[2],
                "sum_dr": Decimal("0"),
                "sum_cr": Decimal("0"),
                "count_dr": 0,
                "count_cr": 0,
            }
        bucket = grouped[key]
        dr = Decimal(str(row["dr"]))
        cr = Decimal(str(row["cr"]))
        bucket["sum_dr"] += dr
        bucket["sum_cr"] += cr
        if dr > 0:
            bucket["count_dr"] += 1
        if cr > 0:
            bucket["count_cr"] += 1

    out = []
    for key in sorted(grouped.keys()):
        bucket = grouped[key]
        out.append(
            {
                "month_key": bucket["month_key"],
                "category": bucket["category"],
                "txn_type": bucket["txn_type"],
                "sum_dr": float(bucket["sum_dr"]),
                "sum_cr": float(bucket["sum_cr"]),
                "count_dr": bucket["count_dr"],
                "count_cr": bucket["count_cr"],
            }
        )
    return out


def _continuity_failures(rows: List[Dict[str, Any]]) -> int:
    ordered = [r for r in rows if r.get("balance") is not None]
    ordered.sort(key=lambda r: (r["txn_date"], r["row_index"]))
    failures = 0
    prev_balance: Optional[Decimal] = None
    for row in ordered:
        balance = Decimal(str(row["balance"])) if row.get("balance") is not None else None
        if balance is None:
            continue
        if prev_balance is not None:
            expected = prev_balance + Decimal(str(row["cr"])) - Decimal(str(row["dr"]))
            if abs(balance - expected) > Decimal("0.01"):
                failures += 1
        prev_balance = balance
    return failures


def _choose_template_sheets(template_path: str) -> Tuple[List[str], List[str]]:
    wb = load_workbook(template_path, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    xns_templates = [s for s in sheet_names if s.upper().startswith("XNS-")]
    pivot_templates = [s for s in sheet_names if s.upper().startswith("PIVOT-")]
    if not xns_templates:
        raise ValueError("No XNS-* template sheet found in workbook")
    if not pivot_templates:
        raise ValueError("No PIVOT-* template sheet found in workbook")
    return xns_templates, pivot_templates


def _upsert_storage_file(path: str, local_file_path: str, content_type: str) -> None:
    bucket = sb.storage.from_(settings.bucket)
    try:
        bucket.remove([path])
    except Exception:
        pass
    with open(local_file_path, "rb") as f:
        bucket.upload(path, f, {"content-type": content_type, "upsert": "true"})


def _to_inr_compact(value: Decimal) -> str:
    n = float(value)
    abs_n = abs(n)
    if abs_n >= 10_000_000:
        return f"₹{n / 10_000_000:.2f} Cr"
    if abs_n >= 100_000:
        return f"₹{n / 100_000:.2f} L"
    return f"₹{n:,.2f}"


def _safe_decimal(value: Any) -> Decimal:
    try:
        return Decimal(str(value))
    except Exception:
        return Decimal("0")


def _load_finance_tag_config() -> FinanceTagConfig:
    cfg: FinanceTagConfig = FinanceTagConfig()
    cfg["pvt_keywords"] = dict(DEFAULT_PVT_KEYWORDS)
    cfg["bank_keywords"] = dict(DEFAULT_BANK_KEYWORDS)
    cfg["false_patterns"] = set(DEFAULT_FALSE_POSITIVES)
    cfg["pvt_entities"] = set()
    cfg["bank_entities"] = set()
    cfg["thresholds"] = dict(DEFAULT_TAG_THRESHOLDS)

    for row in _safe_table_select("finance_keywords", select="domain,keyword,weight,is_active", limit=5000):
        if not row.get("is_active", True):
            continue
        domain = str(row.get("domain") or "").upper().strip()
        keyword = _normalize_text(row.get("keyword") or "")
        if not keyword:
            continue
        weight = _safe_decimal(row.get("weight") or "1")
        if domain == "PVT":
            cfg["pvt_keywords"][keyword] = weight
        elif domain == "BANK":
            cfg["bank_keywords"][keyword] = weight

    for row in _safe_table_select("false_positive_patterns", select="pattern,is_active", limit=1000):
        if row.get("is_active", True):
            pattern = _normalize_text(row.get("pattern") or "")
            if pattern:
                cfg["false_patterns"].add(pattern)

    for row in _safe_table_select("pvt_fin_entities", select="entity_name,aliases,is_active", limit=2000):
        if not row.get("is_active", True):
            continue
        name = _normalize_text(row.get("entity_name") or "")
        if name:
            cfg["pvt_entities"].add(name)
        for alias in row.get("aliases") or []:
            alias_norm = _normalize_text(alias)
            if alias_norm:
                cfg["pvt_entities"].add(alias_norm)

    for row in _safe_table_select("bank_fin_entities", select="entity_name,aliases,is_active", limit=2000):
        if not row.get("is_active", True):
            continue
        name = _normalize_text(row.get("entity_name") or "")
        if name:
            cfg["bank_entities"].add(name)
        for alias in row.get("aliases") or []:
            alias_norm = _normalize_text(alias)
            if alias_norm:
                cfg["bank_entities"].add(alias_norm)

    tag_rows = _safe_table_select("finance_tag_config", select="key,value_json", eq={"key": "thresholds"}, limit=1)
    if tag_rows:
        value = tag_rows[0].get("value_json") or {}
        for key in DEFAULT_TAG_THRESHOLDS.keys():
            if key in value and value[key] is not None:
                if key.endswith("_days") or key.endswith("_hits"):
                    cfg["thresholds"][key] = int(value[key])
                else:
                    cfg["thresholds"][key] = _safe_decimal(value[key])

    return cfg


def _score_keyword_hits(text: str, keyword_weights: Dict[str, Decimal]) -> Tuple[Decimal, List[str]]:
    score = Decimal("0")
    reasons: List[str] = []
    for keyword, weight in keyword_weights.items():
        if keyword and keyword in text:
            score += weight
            reasons.append(f"KW:{keyword}")
    return score, reasons


def _within_days(a: dt.date, b: dt.date, window: int) -> bool:
    return abs((a - b).days) <= window


def _apply_finance_tags(rows: List[Dict[str, Any]], config: FinanceTagConfig) -> List[Dict[str, Any]]:
    if not rows:
        return rows

    threshold_pvt = config["thresholds"]["pvt_min_score"]
    threshold_bank = config["thresholds"]["bank_min_score"]
    weekly_window_days = int(config["thresholds"]["weekly_window_days"])
    weekly_min_hits = int(config["thresholds"]["weekly_min_hits"])
    same_day_split_min_hits = int(config["thresholds"]["same_day_split_min_hits"])
    small_ticket_max = config["thresholds"]["small_ticket_max"]

    by_counterparty_dates: Dict[str, List[dt.date]] = defaultdict(list)
    by_counterparty_day_counts: Dict[Tuple[str, dt.date], int] = Counter()
    by_counterparty_small_tickets: Counter[str] = Counter()

    for row in rows:
        cp = str(row.get("counterparty_norm") or "UNKNOWN")
        d = row.get("txn_date")
        amt = _safe_decimal(row.get("amount") or 0)
        if row.get("dr", 0) > 0 and isinstance(d, dt.date):
            by_counterparty_dates[cp].append(d)
            by_counterparty_day_counts[(cp, d)] += 1
            if amt > 0 and amt <= small_ticket_max:
                by_counterparty_small_tickets[cp] += 1

    sorted_rows = sorted(rows, key=lambda r: (r["txn_date"], r["row_index"]))

    for row in sorted_rows:
        text = _normalize_text(row.get("narration") or "")
        cp = str(row.get("counterparty_norm") or "UNKNOWN")
        amount = _safe_decimal(row.get("amount") or 0)
        txn_date = row.get("txn_date")

        pvt_score = Decimal("0")
        bank_score = Decimal("0")
        reasons: List[str] = []

        if any(pattern and pattern in text for pattern in config["false_patterns"]):
            row["finance_tag"] = None
            row["tag_confidence"] = 0.0
            row["tag_reason_codes"] = ["FALSE_POSITIVE_PATTERN"]
            continue

        # Entity signals
        if any(entity and entity in text for entity in config["pvt_entities"]):
            pvt_score += Decimal("1.40")
            reasons.append("PVT_ENTITY")
        if any(entity and entity in text for entity in config["bank_entities"]):
            bank_score += Decimal("1.55")
            reasons.append("BANK_ENTITY")

        # Keyword scores
        s, kw_reasons = _score_keyword_hits(text, config["pvt_keywords"])
        pvt_score += s
        reasons.extend([f"PVT_{r}" for r in kw_reasons])

        s, kw_reasons = _score_keyword_hits(text, config["bank_keywords"])
        bank_score += s
        reasons.extend([f"BANK_{r}" for r in kw_reasons])

        # Cadence/repetition signals for pvt
        if isinstance(txn_date, dt.date):
            cp_dates = sorted(by_counterparty_dates.get(cp, []))
            near_hits = sum(1 for d in cp_dates if _within_days(d, txn_date, weekly_window_days))
            if near_hits >= weekly_min_hits:
                pvt_score += Decimal("0.65")
                reasons.append("REPEAT_30D")

            weekly_gap_hits = 0
            for i in range(1, len(cp_dates)):
                gap = (cp_dates[i] - cp_dates[i - 1]).days
                if 6 <= gap <= 9 or 13 <= gap <= 16:
                    weekly_gap_hits += 1
            if weekly_gap_hits >= max(1, weekly_min_hits - 1):
                pvt_score += Decimal("0.70")
                reasons.append("WEEKLY_CADENCE")

            if by_counterparty_day_counts[(cp, txn_date)] >= same_day_split_min_hits:
                pvt_score += Decimal("0.45")
                reasons.append("SAME_DAY_SPLIT")

        if by_counterparty_small_tickets[cp] >= 4 and amount <= small_ticket_max:
            pvt_score += Decimal("0.45")
            reasons.append("HF_SMALL_TICKET")

        # Bank EMI regularity signals
        if "EMI" in text or "ECS" in text or "NACH" in text or "ACH" in text:
            bank_score += Decimal("0.55")
            reasons.append("EMI_PATTERN")

        # Disbursal/inflow patterns for bank financing
        if row.get("cr", 0) > 0 and ("DISBURS" in text or "LOAN" in text):
            bank_score += Decimal("0.50")
            reasons.append("BANK_DISBURSAL")

        tag: Optional[str] = None
        score_used = Decimal("0")

        if bank_score >= threshold_bank:
            tag = "BANK_FIN"
            score_used = bank_score
        if pvt_score >= threshold_pvt and tag is None:
            tag = "PVT_FIN"
            score_used = pvt_score
        if pvt_score >= threshold_pvt and bank_score >= threshold_bank:
            # Priority: BANK over PVT
            tag = "BANK_FIN"
            score_used = bank_score
            reasons.append("BANK_OVERRIDE")

        conf_den = max(threshold_bank, threshold_pvt) * Decimal("1.8")
        confidence = float(min(Decimal("1"), max(Decimal("0"), score_used / conf_den))) if tag else 0.0

        row["finance_tag"] = tag
        row["tag_confidence"] = round(confidence, 5)
        row["tag_reason_codes"] = sorted(set(reasons))

    return rows


def _compute_risk_summary(rows: Sequence[Dict[str, Any]]) -> Dict[str, Any]:
    total_debits = sum(_safe_decimal(r.get("dr") or 0) for r in rows)
    pvt_debits = sum(_safe_decimal(r.get("dr") or 0) for r in rows if r.get("finance_tag") == "PVT_FIN")
    bank_debits = sum(_safe_decimal(r.get("dr") or 0) for r in rows if r.get("finance_tag") == "BANK_FIN")

    pvt_share = float((pvt_debits / total_debits) * Decimal("100")) if total_debits > 0 else 0.0
    bank_share = float((bank_debits / total_debits) * Decimal("100")) if total_debits > 0 else 0.0

    pvt_by_cp: Dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    for r in rows:
        if r.get("finance_tag") == "PVT_FIN":
            pvt_by_cp[str(r.get("counterparty_norm") or "UNKNOWN")] += _safe_decimal(r.get("dr") or 0)

    top_cp_share = 0.0
    if pvt_by_cp and pvt_debits > 0:
        top_cp_share = float(max(pvt_by_cp.values()) / pvt_debits)

    weekly_repetition = any("WEEKLY_CADENCE" in (r.get("tag_reason_codes") or []) for r in rows)
    emi_miss = any(
        (r.get("finance_tag") == "BANK_FIN")
        and ("RETURN" in str(r.get("category") or "").upper() or "BOUNCE" in _normalize_text(r.get("narration") or ""))
        for r in rows
    )
    bank_emi_counterparties = {
        str(r.get("counterparty_norm") or "UNKNOWN")
        for r in rows
        if r.get("finance_tag") == "BANK_FIN" and _safe_decimal(r.get("dr") or 0) > 0
    }
    multiple_emis = len(bank_emi_counterparties) >= 4

    score = 0
    reasons: List[str] = []

    if pvt_share >= 50:
        score += 35
        reasons.append("PVT exposure >=50% of debits")
    elif pvt_share >= 30:
        score += 20
        reasons.append("PVT exposure >=30% of debits")

    if top_cp_share >= 0.5:
        score += 10
        reasons.append("Top private lender concentration is high")

    if weekly_repetition:
        score += 10
        reasons.append("Weekly / biweekly repayment cadence detected")

    if emi_miss:
        score += 15
        reasons.append("EMI/bounce miss signals detected")

    if multiple_emis:
        score += 10
        reasons.append("Multiple EMI counterparties detected")

    if pvt_share < 10:
        score -= 10
        reasons.append("Low private-lender share (<10%)")

    if bank_debits > 0 and not emi_miss:
        score -= 10
        reasons.append("Stable EMI behavior")

    score = max(0, min(100, score))

    if score <= 24:
        band = "Low"
    elif score <= 49:
        band = "Medium"
    elif score <= 74:
        band = "High"
    else:
        band = "Very High"

    return {
        "risk_score": score,
        "risk_band": band,
        "reasons": reasons,
        "pvt_share_of_debits": round(pvt_share, 2),
        "bank_share_of_debits": round(bank_share, 2),
        "top_lender_concentration": round(top_cp_share * 100, 2),
        "weekly_repetition": weekly_repetition,
        "emi_miss": emi_miss,
        "multiple_emis": multiple_emis,
    }


def _update_version(version_id: str, payload: Dict[str, Any]) -> None:
    """Update with compatibility fallback for older schemas."""
    try:
        sb.table("statement_versions").update(payload).eq("id", version_id).execute()
        return
    except Exception:
        pass

    legacy_keys = {
        k: v
        for k, v in payload.items()
        if k in {"status", "run_at", "unmapped_txn_lines", "continuity_failures", "excel_url", "parse_hash"}
    }
    if legacy_keys:
        sb.table("statement_versions").update(legacy_keys).eq("id", version_id).execute()


@app.get("/health")
def health() -> Dict[str, Any]:
    template_exists = Path(settings.template_path).exists()
    workbook_enabled = _env_flag("STATEMENT_WORKBOOK_ENABLED", True)
    workbook_active = workbook_enabled and template_exists
    return {
        "ok": True,
        "template_path": settings.template_path,
        "template_exists": template_exists,
        "workbook_enabled": workbook_enabled,
        "workbook_active": workbook_active,
        "bucket": settings.bucket,
    }


@app.post("/jobs/parse_statement/{version_id}")
def parse_statement(version_id: str, force: bool = False) -> Dict[str, Any]:
    now = dt.datetime.now(dt.timezone.utc)
    now_iso = now.isoformat()
    template_exists = Path(settings.template_path).exists()
    workbook_enabled = _env_flag("STATEMENT_WORKBOOK_ENABLED", True)
    workbook_active = workbook_enabled and template_exists
    workbook_skip_reason: Optional[str] = None
    if not workbook_enabled:
        workbook_skip_reason = "Workbook generation disabled by STATEMENT_WORKBOOK_ENABLED"
    elif not template_exists:
        workbook_skip_reason = f"Workbook template not found: {settings.template_path}"

    version_rows = _safe_table_select("statement_versions", select="*", eq={"id": version_id}, limit=1)
    if not version_rows:
        raise HTTPException(status_code=404, detail="Statement version not found")
    version_row = version_rows[0]

    statement_id = version_row.get("statement_id")
    statement_rows = _safe_table_select("statements", select="id,lead_id", eq={"id": statement_id}, limit=1)
    statement_row = statement_rows[0] if statement_rows else {"id": statement_id, "lead_id": None}

    pdf_resp = sb.table("pdf_files").select("*").eq("version_id", version_id).order("created_at").execute()
    pdfs = pdf_resp.data or []
    if not pdfs:
        raise HTTPException(status_code=404, detail="No PDFs found for this version")

    parse_hash = _hash_uid(
        [version_id]
        + [
            f"{p.get('id')}|{p.get('storage_path')}|{p.get('original_name') or p.get('file_name') or ''}|{p.get('created_at') or ''}"
            for p in pdfs
        ]
    )

    if (
        not force
        and str(version_row.get("parse_hash") or "") == parse_hash
        and str(version_row.get("parse_status") or "").upper() == "SUCCESS"
        and (
            not workbook_active
            or (version_row.get("underwriting_workbook_url") or version_row.get("excel_url"))
        )
    ):
        return {
            "status": "READY",
            "idempotent": True,
            "version_id": version_id,
            "parse_hash": parse_hash,
            "excel_path": version_row.get("excel_url"),
            "workbook_path": version_row.get("underwriting_workbook_url") or version_row.get("excel_url"),
            "workbook_enabled": workbook_enabled,
            "workbook_active": workbook_active,
            "workbook_skip_reason": workbook_skip_reason,
        }

    _update_version(
        version_id,
        {
            "status": "PARSING",
            "parse_status": "RUNNING",
            "parse_started_at": now_iso,
            "parse_completed_at": None,
            "error_reason": None,
            "raw_row_count": 0,
            "parsed_row_count": 0,
            "run_at": now_iso,
            "parse_hash": parse_hash,
            "unmapped_txn_lines": 0,
            "continuity_failures": 0,
        },
    )

    try:
        # Keep re-runs deterministic and idempotent.
        for table in [
            "raw_statement_lines",
            "transactions",
            "aggregates_monthly",
            "pivots",
            "statement_transaction_ledger",
        ]:
            try:
                sb.table(table).delete().eq("version_id", version_id).execute()
            except Exception:
                pass

        tmp_dir = tempfile.mkdtemp(prefix=f"stmt_{version_id}_")
        tag_cfg = _load_finance_tag_config()

        raw_lines_all: List[Tuple[Dict[str, Any], List[Any], List[Dict[str, Any]]]] = []
        all_raw_insert_rows: List[Dict[str, Any]] = []

        for pdf in pdfs:
            storage_path = pdf.get("storage_path")
            if not storage_path:
                raise HTTPException(status_code=400, detail=f"PDF {pdf.get('id')} missing storage_path")

            binary = sb.storage.from_(settings.bucket).download(storage_path)
            local_pdf = os.path.join(tmp_dir, f"{pdf['id']}.pdf")
            with open(local_pdf, "wb") as f:
                f.write(binary)

            raw_lines = extract_raw_lines_pdfplumber(local_pdf)
            raw_insert_rows: List[Dict[str, Any]] = []
            for line in raw_lines:
                raw_id = str(uuid.uuid4())
                row = {
                    "id": raw_id,
                    "version_id": version_id,
                    "pdf_file_id": pdf["id"],
                    "page_no": line.page_no,
                    "row_no": line.row_no,
                    "raw_row_text": line.raw_row_text,
                    "raw_date_text": line.date_text,
                    "raw_narration_text": line.narration_text,
                    "raw_dr_text": line.dr_text,
                    "raw_cr_text": line.cr_text,
                    "raw_balance_text": line.bal_text,
                    "line_type": line.line_type,
                    "extraction_method": line.extraction_method,
                    "bbox_json": None,
                }
                raw_insert_rows.append(row)
                all_raw_insert_rows.append(row)

            raw_lines_all.append((pdf, raw_lines, raw_insert_rows))

        _batch_insert("raw_statement_lines", all_raw_insert_rows, size=1000)

        mapped_indices_by_pdf: Dict[str, Set[int]] = defaultdict(set)
        transactions_to_insert: List[Dict[str, Any]] = []
        excel_txns_by_pdf: Dict[str, List[Dict[str, Any]]] = defaultdict(list)

        raw_txn_candidate_count = 0
        raw_dr_total = Decimal("0")
        raw_cr_total = Decimal("0")

        row_index_global = 0

        for pdf, raw_lines, inserted_rows in raw_lines_all:
            merged = merge_multiline_transactions(raw_lines)
            for merged_row in merged:
                dr, cr, bal = _infer_amount_triplet(
                    merged_row.get("dr_text"),
                    merged_row.get("cr_text"),
                    merged_row.get("bal_text"),
                    merged_row.get("narration") or merged_row.get("raw_row_text") or "",
                )
                raw_txn_candidate_count += 1
                raw_dr_total += dr
                raw_cr_total += cr

                txn_date = _parse_date(merged_row.get("date_text", ""))
                if txn_date is None:
                    continue

                raw_indices = merged_row["raw_indices"]
                raw_ids = [inserted_rows[i]["id"] for i in raw_indices if 0 <= i < len(inserted_rows)]
                if not raw_ids:
                    continue

                for index in raw_indices:
                    mapped_indices_by_pdf[pdf["id"]].add(index)

                row_index_global += 1
                narration = (merged_row.get("narration") or "").strip() or "-"
                category = _classify_txn_legacy(narration, dr, cr)
                txn_type = _txn_type(dr, cr)
                amount = max(abs(dr), abs(cr))
                counterparty = _normalize_counterparty(narration)

                dedupe_hash = _hash_uid(
                    [
                        statement_id,
                        txn_date.isoformat(),
                        f"{amount:.2f}",
                        narration,
                        f"{(bal if bal is not None else Decimal('0')):.2f}",
                        row_index_global,
                    ]
                )

                transaction_uid = _hash_uid([version_id, dedupe_hash])
                tx_row = {
                    "id": str(uuid.uuid4()),
                    "version_id": version_id,
                    "raw_line_ids": raw_ids,
                    "txn_date": txn_date,
                    "month_key": _month_key(txn_date),
                    "narration": narration,
                    "dr": float(dr),
                    "cr": float(cr),
                    "balance": float(bal) if bal is not None else None,
                    "counterparty_norm": counterparty,
                    "txn_type": txn_type,
                    "category": category,
                    "flags": [],
                    "transaction_uid": transaction_uid,
                    "row_index": row_index_global,
                    "amount": float(amount),
                    "dedupe_hash": dedupe_hash,
                    "pdf_file_id": pdf["id"],
                    "raw_indices": raw_indices,
                    "raw_json": {
                        "pdf_file_id": pdf["id"],
                        "raw_indices": raw_indices,
                        "date_text": merged_row.get("date_text"),
                        "dr_text": merged_row.get("dr_text"),
                        "cr_text": merged_row.get("cr_text"),
                        "bal_text": merged_row.get("bal_text"),
                    },
                }
                transactions_to_insert.append(tx_row)

        transactions_to_insert = _apply_finance_tags(transactions_to_insert, tag_cfg)

        for tx in transactions_to_insert:
            txn_date = tx["txn_date"]
            excel_txns_by_pdf[tx["pdf_file_id"]].append(
                {
                    "date": txn_date,
                    "date_label": _to_date_label(txn_date),
                    "month_label": _month_label(txn_date),
                    "txn_type": tx.get("txn_type", ""),
                    "ref_no": "",
                    "category": tx.get("category", ""),
                    "narration": tx.get("narration", ""),
                    "dr": float(tx.get("dr") or 0),
                    "cr": float(tx.get("cr") or 0),
                    "balance": float(tx.get("balance") or 0),
                    "finance_tag": tx.get("finance_tag"),
                    "tag_confidence": tx.get("tag_confidence") or 0.0,
                    "reason_codes": ", ".join(tx.get("tag_reason_codes") or []),
                }
            )

        parsed_row_count = len(transactions_to_insert)
        parsed_dr_total = sum(_safe_decimal(tx.get("dr") or 0) for tx in transactions_to_insert)
        parsed_cr_total = sum(_safe_decimal(tx.get("cr") or 0) for tx in transactions_to_insert)

        unmapped_total = 0
        for pdf, raw_lines, _ in raw_lines_all:
            unmapped_total += reconcile_strict(raw_lines, mapped_indices_by_pdf[pdf["id"]])

        strict_error_reasons: List[str] = []
        if unmapped_total > 0:
            strict_error_reasons.append(f"UNMAPPED_TRANSACTION_LINES:{unmapped_total}")
        if raw_txn_candidate_count != parsed_row_count:
            strict_error_reasons.append(
                f"ROW_COUNT_MISMATCH:raw={raw_txn_candidate_count},parsed={parsed_row_count}"
            )
        if abs(raw_dr_total - parsed_dr_total) > Decimal("0.01") or abs(raw_cr_total - parsed_cr_total) > Decimal("0.01"):
            strict_error_reasons.append(
                f"TOTAL_MISMATCH:raw_dr={raw_dr_total},parsed_dr={parsed_dr_total},raw_cr={raw_cr_total},parsed_cr={parsed_cr_total}"
            )

        if strict_error_reasons:
            _update_version(
                version_id,
                {
                    "status": "PARSE_FAILED",
                    "parse_status": "FAILED",
                    "error_reason": "; ".join(strict_error_reasons),
                    "unmapped_txn_lines": unmapped_total,
                    "raw_row_count": raw_txn_candidate_count,
                    "parsed_row_count": parsed_row_count,
                    "parse_completed_at": now_iso,
                    "run_at": now_iso,
                },
            )
            return {
                "status": "PARSE_FAILED",
                "version_id": version_id,
                "parse_hash": parse_hash,
                "reasons": strict_error_reasons,
                "unmapped": unmapped_total,
                "raw_row_count": raw_txn_candidate_count,
                "parsed_row_count": parsed_row_count,
            }

        _batch_insert(
            "transactions",
            [
                {
                    "id": tx["id"],
                    "version_id": version_id,
                    "raw_line_ids": tx["raw_line_ids"],
                    "txn_date": tx["txn_date"].isoformat(),
                    "month_key": tx["month_key"],
                    "narration": tx["narration"],
                    "dr": tx["dr"],
                    "cr": tx["cr"],
                    "balance": tx["balance"],
                    "counterparty_norm": tx["counterparty_norm"],
                    "txn_type": tx["txn_type"],
                    "category": tx["category"],
                    "flags": tx["flags"],
                    "transaction_uid": tx["transaction_uid"],
                    "finance_tag": tx.get("finance_tag"),
                    "tag_confidence": tx.get("tag_confidence"),
                    "tag_reason_codes": tx.get("tag_reason_codes") or [],
                }
                for tx in transactions_to_insert
            ],
            size=500,
        )

        # Source-of-truth ledger (strict dedupe + raw row capture)
        ledger_rows = [
            {
                "statement_id": statement_id,
                "version_id": version_id,
                "account_id": None,
                "txn_id": tx["transaction_uid"],
                "txn_date": tx["txn_date"].isoformat(),
                "narration": tx["narration"],
                "dr": tx["dr"],
                "cr": tx["cr"],
                "amount": tx["amount"],
                "balance": tx["balance"],
                "raw_row_json": tx["raw_json"],
                "row_index": tx["row_index"],
                "finance_tag": tx.get("finance_tag"),
                "tag_confidence": tx.get("tag_confidence"),
                "tag_reason_codes": tx.get("tag_reason_codes") or [],
                "dedupe_hash": tx["dedupe_hash"],
            }
            for tx in transactions_to_insert
        ]
        try:
            _batch_insert("statement_transaction_ledger", ledger_rows, size=500)
        except Exception:
            # Keep service backwards-compatible when ledger table not yet migrated.
            pass

        monthly_aggregates = _build_monthly_aggregates(transactions_to_insert)
        _batch_insert(
            "aggregates_monthly",
            [
                {
                    "id": str(uuid.uuid4()),
                    "version_id": version_id,
                    "month_key": agg["month_key"],
                    "kpis": agg["kpis"],
                }
                for agg in monthly_aggregates
            ],
            size=200,
        )

        pivot_rows = _build_pivot_rows(transactions_to_insert)
        _batch_insert(
            "pivots",
            [
                {
                    "id": str(uuid.uuid4()),
                    "version_id": version_id,
                    "month_key": p["month_key"],
                    "category": p["category"],
                    "txn_type": p["txn_type"],
                    "sum_dr": p["sum_dr"],
                    "sum_cr": p["sum_cr"],
                    "count_dr": p["count_dr"],
                    "count_cr": p["count_cr"],
                }
                for p in pivot_rows
            ],
            size=500,
        )

        continuity_failures = _continuity_failures(transactions_to_insert)
        risk = _compute_risk_summary(transactions_to_insert)

        pvt_fin_rows = [
            {
                "date": tx["txn_date"],
                "month_label": _month_label(tx["txn_date"]),
                "type": "PVT FIN",
                "category": tx.get("category") or "",
                "dr": tx.get("dr") or 0,
                "cr": tx.get("cr") or 0,
                "narration": tx.get("narration") or "",
            }
            for tx in transactions_to_insert
            if tx.get("finance_tag") == "PVT_FIN"
        ]

        bank_fin_rows = [
            {
                "date": tx["txn_date"],
                "month_label": _month_label(tx["txn_date"]),
                "type": "BANK FIN",
                "category": tx.get("category") or "",
                "dr": tx.get("dr") or 0,
                "cr": tx.get("cr") or 0,
                "narration": tx.get("narration") or "",
            }
            for tx in transactions_to_insert
            if tx.get("finance_tag") == "BANK_FIN"
        ]

        cons_rows = []
        for m in sorted(monthly_aggregates, key=lambda x: x["month_key"]):
            kpis = m["kpis"]
            cons_rows.append(
                {
                    "month_key": m["month_key"],
                    "total_dr": float(kpis.get("debit_total") or 0),
                    "total_cr": float(kpis.get("credit_total") or 0),
                    "net": float((kpis.get("credit_total") or 0) - (kpis.get("debit_total") or 0)),
                }
            )

        pvt_debit_total = sum(
            _safe_decimal(x.get("dr") or 0)
            for x in transactions_to_insert
            if x.get("finance_tag") == "PVT_FIN"
        )
        bank_debit_total = sum(
            _safe_decimal(x.get("dr") or 0)
            for x in transactions_to_insert
            if x.get("finance_tag") == "BANK_FIN"
        )

        analysis_rows = [
            ["Parse Hash", parse_hash],
            ["Raw Transaction Rows", raw_txn_candidate_count],
            ["Parsed Transaction Rows", parsed_row_count],
            ["Total Debits", float(parsed_dr_total)],
            ["Total Credits", float(parsed_cr_total)],
            ["PVT Share of Debits (%)", risk["pvt_share_of_debits"]],
            ["BANK Share of Debits (%)", risk["bank_share_of_debits"]],
            ["Top Lender Concentration (%)", risk["top_lender_concentration"]],
            ["Risk Score", risk["risk_score"]],
            ["Risk Band", risk["risk_band"]],
            [
                "Exposure Summary",
                f"PVT {_to_inr_compact(pvt_debit_total)} | BANK {_to_inr_compact(bank_debit_total)}",
            ],
        ]

        final_rows = [["Risk Band", risk["risk_band"]], ["Risk Score", risk["risk_score"]]]
        for i, reason in enumerate(risk["reasons"], start=1):
            final_rows.append([f"Reason {i}", reason])

        legacy_excel_path: Optional[str] = None
        workbook_path: Optional[str] = None
        workbook_generated_at: Optional[str] = None

        if workbook_active:
            xns_templates, pivot_templates = _choose_template_sheets(settings.template_path)
            accounts = []
            for index, pdf in enumerate(pdfs):
                txns = excel_txns_by_pdf.get(pdf["id"], [])
                xns_tpl = xns_templates[min(index, len(xns_templates) - 1)]
                piv_tpl = pivot_templates[min(index, len(pivot_templates) - 1)]
                accounts.append(
                    {
                        "xns_template_sheet": xns_tpl,
                        "pivot_template_sheet": piv_tpl,
                        "xns_sheet_name": f"XNS-{index + 1}",
                        "pivot_sheet_name": f"PIVOT-{index + 1}",
                        "xns_start_row": 10,
                        "xns_template_row": 10,
                        "pivot_start_row": 2,
                        "pivot_template_row": 2,
                        "txns": txns,
                        "pivots": [p for p in pivot_rows],
                    }
                )

            out_xlsx = os.path.join(tmp_dir, "underwriting_workbook.xlsx")
            generate_perfios_excel(
                template_path=settings.template_path,
                output_path=out_xlsx,
                context={
                    "accounts": accounts,
                    "analysis_rows": analysis_rows,
                    "analysis_start_row": 2,
                    "cons_rows": cons_rows,
                    "pvt_fin_rows": pvt_fin_rows,
                    "bank_fin_rows": bank_fin_rows,
                    "final_rows": final_rows,
                },
            )

            legacy_excel_path = f"exports/{version_id}/perfios_output.xlsx"
            _upsert_storage_file(
                legacy_excel_path,
                out_xlsx,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            lead_id = statement_row.get("lead_id") or "unknown"
            workbook_path = f"underwriting/{lead_id}/{statement_id}/underwriting_workbook.xlsx"
            _upsert_storage_file(
                workbook_path,
                out_xlsx,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            try:
                _batch_upsert(
                    "statement_underwriting_workbooks",
                    [
                        {
                            "lead_id": statement_row.get("lead_id"),
                            "statement_id": statement_id,
                            "version_id": version_id,
                            "parse_hash": parse_hash,
                            "storage_path": workbook_path,
                            "meta_json": {
                                "raw_row_count": raw_txn_candidate_count,
                                "parsed_row_count": parsed_row_count,
                                "risk_score": risk["risk_score"],
                                "risk_band": risk["risk_band"],
                            },
                        }
                    ],
                    on_conflict="version_id,parse_hash",
                    size=100,
                )
            except Exception:
                pass

            workbook_generated_at = now_iso

        _update_version(
            version_id,
            {
                "status": "READY",
                "parse_status": "SUCCESS",
                "error_reason": None,
                "excel_url": legacy_excel_path,
                "underwriting_workbook_url": workbook_path,
                "underwriting_workbook_generated_at": workbook_generated_at,
                "unmapped_txn_lines": 0,
                "continuity_failures": continuity_failures,
                "raw_row_count": raw_txn_candidate_count,
                "parsed_row_count": parsed_row_count,
                "parse_completed_at": now_iso,
                "run_at": now_iso,
                "parse_hash": parse_hash,
            },
        )

        try:
            sb.table("audit_events").insert(
                {
                    "id": str(uuid.uuid4()),
                    "entity_type": "statement_version",
                    "entity_id": version_id,
                    "action": "PARSE_READY",
                    "actor_user_id": None,
                    "payload": {
                        "pdf_count": len(pdfs),
                        "transactions": parsed_row_count,
                        "continuity_failures": continuity_failures,
                        "excel_url": legacy_excel_path,
                        "workbook_url": workbook_path,
                        "workbook_enabled": workbook_enabled,
                        "workbook_active": workbook_active,
                        "workbook_skip_reason": workbook_skip_reason,
                        "parse_hash": parse_hash,
                        "risk_score": risk["risk_score"],
                        "risk_band": risk["risk_band"],
                    },
                }
            ).execute()
        except Exception:
            pass

        return {
            "status": "READY",
            "version_id": version_id,
            "parse_hash": parse_hash,
            "excel_path": legacy_excel_path,
            "workbook_path": workbook_path,
            "workbook_enabled": workbook_enabled,
            "workbook_active": workbook_active,
            "workbook_skip_reason": workbook_skip_reason,
            "transactions": parsed_row_count,
            "raw_row_count": raw_txn_candidate_count,
            "parsed_row_count": parsed_row_count,
            "continuity_failures": continuity_failures,
            "risk": risk,
        }
    except HTTPException:
        raise
    except Exception as exc:  # pragma: no cover
        _update_version(
            version_id,
            {
                "status": "PARSE_FAILED",
                "parse_status": "FAILED",
                "error_reason": str(exc),
                "parse_completed_at": now_iso,
                "run_at": now_iso,
            },
        )
        raise HTTPException(status_code=500, detail=f"parse_statement failed: {exc}") from exc
