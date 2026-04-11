from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
if str(Path(__file__).resolve().parent) not in sys.path:
    sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import load_workbook

from verify_workbook_types import read_manual_xns, read_parsed_rows, verify


def default_cases_dir() -> Path:
    return Path.home() / "Downloads"


def load_manifest(path: Path) -> list[dict]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    return list(payload.get("cases", []))


def resolve_manual_workbook(base_dir: Path, relative_or_absolute: str) -> Path:
    candidate = Path(relative_or_absolute).expanduser()
    if candidate.is_absolute():
        return candidate
    return base_dir / candidate


def _as_text(value) -> str:
    return "" if value is None else str(value)


def _export_current_workbook(job_id: str, export_dir: Path) -> Path:
    import app.models  # noqa: F401
    from app.db.session import SessionLocal
    from app.services.export_service import ExportService

    db = SessionLocal()
    try:
        service = ExportService(db, export_dir)
        export = service.export_job_workbook(job_id)
        return Path(export.file_path)
    finally:
        db.close()


def _check_workbook_assertions(workbook_path: Path, assertions: list[dict]) -> list[str]:
    wb = load_workbook(workbook_path, data_only=True)
    failures: list[str] = []
    for assertion in assertions:
        sheet = assertion["sheet"]
        cell = assertion["cell"]
        if sheet not in wb.sheetnames:
            failures.append(f"missing sheet {sheet}")
            continue
        actual = _as_text(wb[sheet][cell].value)
        expected = assertion.get("equals")
        if expected is not None and actual != expected:
            failures.append(f"{sheet}!{cell} expected {expected!r} got {actual!r}")
        contains = assertion.get("contains")
        if contains is not None and contains not in actual:
            failures.append(f"{sheet}!{cell} expected to contain {contains!r} got {actual!r}")
    return failures


def _check_account_thresholds(summary_rows: list[dict[str, str]], expected_accounts: dict[str, dict]) -> list[str]:
    failures: list[str] = []
    summary_by_account = {row["account"]: row for row in summary_rows}
    for account, thresholds in expected_accounts.items():
        row = summary_by_account.get(account)
        if row is None:
            failures.append(f"missing summary row for account {account}")
            continue
        if float(row["match_pct"]) < float(thresholds.get("min_match_pct", 0.0)):
            failures.append(f"{account} match_pct {row['match_pct']} below {thresholds['min_match_pct']}")
        if float(row["type_match_pct"]) < float(thresholds.get("min_type_match_pct", 0.0)):
            failures.append(f"{account} type_match_pct {row['type_match_pct']} below {thresholds['min_type_match_pct']}")
        if int(row["manual_only"]) > int(thresholds.get("max_manual_only", 0)):
            failures.append(f"{account} manual_only {row['manual_only']} above {thresholds['max_manual_only']}")
        if int(row["parsed_only"]) > int(thresholds.get("max_parsed_only", 0)):
            failures.append(f"{account} parsed_only {row['parsed_only']} above {thresholds['max_parsed_only']}")
    return failures


@dataclass
class RegressionResult:
    failures: list[str]
    exports: dict[str, Path]


def run_manifest(manifest_path: Path, db_path: Path, export_dir: Path, cases_dir: Path) -> RegressionResult:
    failures: list[str] = []
    exports: dict[str, Path] = {}

    for case in load_manifest(manifest_path):
        case_name = case["name"]
        manual_workbook = resolve_manual_workbook(cases_dir, case["manual_workbook"])
        if not manual_workbook.exists():
            failures.append(f"{case_name}: manual workbook missing at {manual_workbook}")
            continue

        export_path = _export_current_workbook(case["job_id"], export_dir)
        exports[case_name] = export_path

        workbook_failures = _check_workbook_assertions(export_path, list(case.get("workbook_assertions", [])))
        for failure in workbook_failures:
            failures.append(f"{case_name}: {failure}")

        manual = read_manual_xns(manual_workbook)
        parsed = read_parsed_rows(db_path, case["job_id"])
        summaries, _details = verify(manual, parsed)
        threshold_failures = _check_account_thresholds(summaries, dict(case.get("accounts", {})))
        for failure in threshold_failures:
            failures.append(f"{case_name}: {failure}")

    return RegressionResult(failures=failures, exports=exports)


def main() -> int:
    parser = argparse.ArgumentParser(description="Run manifest-driven real-case regression checks")
    parser.add_argument("--manifest", default="tests/fixtures/case_regressions.json")
    parser.add_argument("--db", default="bank_intel.db")
    parser.add_argument("--export-dir", default="data/exports")
    parser.add_argument("--cases-dir", default=str(default_cases_dir()))
    args = parser.parse_args()

    result = run_manifest(
        manifest_path=Path(args.manifest),
        db_path=Path(args.db),
        export_dir=Path(args.export_dir),
        cases_dir=Path(args.cases_dir).expanduser(),
    )

    for case_name, path in result.exports.items():
        print(f"{case_name}: exported {path}")

    if result.failures:
        print("regression failures:")
        for failure in result.failures:
            print(f" - {failure}")
        return 1

    print("all case regressions passed")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
