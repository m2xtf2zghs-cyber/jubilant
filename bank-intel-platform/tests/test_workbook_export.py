from datetime import date
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

from app.workbook.exporter import write_workbook


class DummyTxn(SimpleNamespace):
    pass


def _txn(**kwargs):
    defaults = dict(
        source_bank="BOI",
        source_account_no="802525130000005",
        source_account_name="DHARA LOGISTICS",
        source_account_type="OD",
        source_file="statement.pdf",
        txn_date=date(2026, 1, 5),
        txn_order=1,
        page_no=1,
        line_ref=1,
        classification_primary="SALES",
        normalized_party="PARTY A",
        cheque_no=None,
        raw_narration="RTGS/ABC",
        cleaned_narration="RTGS/ABC",
        debit=0.0,
        credit=100000.0,
        balance=50000.0,
        month_key="JAN(26)",
        odd_figure_flag=False,
        doubt_flag=False,
        bank_fin_flag=False,
        private_fin_flag=False,
        return_flag=False,
        account=None,
    )
    defaults.update(kwargs)
    return DummyTxn(**defaults)


def test_workbook_contains_required_sheets(tmp_path: Path) -> None:
    txns = {
        "BOI-0005-OD": [
            _txn(classification_primary="SALES", credit=200000),
            _txn(classification_primary="PURCHASE", debit=150000, credit=0),
            _txn(classification_primary="NAMES", debit=12000, credit=0),
            _txn(classification_primary="DOUBT", debit=1000, credit=0, doubt_flag=True),
            _txn(classification_primary="RETURN", debit=2500, credit=0, raw_narration="CHQ RET", cleaned_narration="CHQ RET"),
            _txn(classification_primary="ODD FIG", debit=0, credit=500000, odd_figure_flag=True),
            _txn(classification_primary="BANK FIN", debit=42004, credit=0, bank_fin_flag=True),
            _txn(classification_primary="PVT FIN", debit=20000, credit=0, private_fin_flag=True),
        ]
    }
    out = tmp_path / "out.xlsx"
    write_workbook(out, txns)

    wb = load_workbook(out)
    names = set(wb.sheetnames)
    assert "ANALYSIS" in names
    assert any(n.startswith("XNS-") for n in names)
    assert any(n.startswith("PIVOT-") for n in names)
    assert {"ODD FIG", "DOUBT", "NAMES", "BANK FIN", "PVT FIN", "RETURN", "FINAL"}.issubset(names)


def test_workbook_layout_matches_blueprint_shape(tmp_path: Path) -> None:
    txns = {
        "TMB-5048-CA": [
            _txn(
                source_bank="TMB",
                source_account_no="210150310875048",
                source_account_name="AFFAN METALS",
                source_account_type="CA",
                txn_date=date(2025, 8, 1),
                classification_primary="SALES",
                credit=250000,
                raw_narration="RTGS/METAL AND SC/HDFC",
            ),
            _txn(
                source_bank="TMB",
                source_account_no="210150310875048",
                source_account_name="AFFAN METALS",
                source_account_type="CA",
                txn_date=date(2025, 8, 5),
                classification_primary="ODD FIG",
                credit=500000,
                raw_narration="RTGS/ROUND CREDIT",
                odd_figure_flag=True,
            ),
            _txn(
                source_bank="TMB",
                source_account_no="210150310875048",
                source_account_name="AFFAN METALS",
                source_account_type="CA",
                txn_date=date(2025, 10, 10),
                classification_primary="RETURN",
                debit=409422,
                credit=0,
                raw_narration="I/W RTGS RTN -(NOT REP)",
                cleaned_narration="I/W RTGS RTN -(NOT REP)",
                return_flag=True,
            ),
        ],
        "TMB-0123-OD": [
            _txn(
                source_bank="TMB",
                source_account_no="210150310000123",
                source_account_name="VEERA INDUSTRIES",
                source_account_type="OD",
                txn_date=date(2025, 8, 1),
                classification_primary="PURCHASE",
                debit=175000,
                credit=0,
                raw_narration="RTGS/SUPPLIER",
            ),
            _txn(
                source_bank="TMB",
                source_account_no="210150310000123",
                source_account_name="VEERA INDUSTRIES",
                source_account_type="OD",
                txn_date=date(2025, 8, 2),
                classification_primary="NAMES",
                debit=12000,
                credit=0,
                raw_narration="SALARY/ANAND T",
            ),
        ],
    }
    out = tmp_path / "layout.xlsx"
    write_workbook(
        out,
        txns,
        case_context={"case_metadata": {"analyst_name": "CODEX", "other_bank_note": "YES"}},
    )

    wb = load_workbook(out, data_only=True)

    odd_fig = wb["ODD FIG"]
    assert odd_fig["C1"].value == "TMB-5048-CA"
    assert odd_fig["A2"].value == "Date"
    assert odd_fig["C3"].value == "ODD FIG"

    returns = wb["RETURN"]
    assert returns["C1"].value == "TMB-5048-CA"
    assert returns["C3"].value == "RETURN"

    pivot = wb["PIVOT-TMB-5048-CA"]
    assert pivot["D1"].value == "AFFAN METALS\nTMB-5048-CA"
    assert pivot["A5"].value == "ANALYSED BY : CODEX"
    assert pivot["C6"].value == "MONTH"
    assert pivot["D6"].value == "Values"
    assert pivot["C7"].value == "AUG"

    cons = wb["CONS"]
    assert "AFFAN METALS & VEERA INDUSTRIES" in cons["D2"].value
    assert cons["D5"].value == "MONTH"
    assert cons["E5"].value == "AFFAN METALS\nTMB-5048-CA\nPURCHASE"
    assert cons["H5"].value == "AFFAN METALS\nTMB-5048-CA\nSALES"

    final = wb["FINAL"]
    assert final["G14"].value == "INWARD CHEQUE RETURN"
    assert final["G16"].value == "IRREUGULAR REPAYMENT FINANCE"

    xns = wb["XNS-TMB-5048-CA"]
    assert xns["A1"].value == "Sl. No."
    assert xns["D2"].value == "SALES"
