from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any

from app.utils.money import to_lakhs

if TYPE_CHECKING:
    from app.models.transaction import Transaction
else:  # pragma: no cover
    Transaction = Any


TYPE_COLOR = {
    "UNMATCH INB TRF": "FFA500",
    "SIS CON": "6699FF",
    "PURCHASE": "FFFF00",
    "SALES": "90EE90",
}

PIVOT_TYPE_ORDER = [
    "PROP",
    "UNMATCH INB TRF",
    "INB TRF",
    "SIS CON",
    "UNMATCH SIS CON",
    "ODD FIG",
    "DOUBT",
    "CASH",
    "NAMES",
    "REVERSAL",
    "RETURN",
    "INSURANCE",
    "SALES",
    "PURCHASE",
    "EXPENSE",
    "BANK FIN",
    "PVT FIN",
]

MONTH_ORDER = {
    "JAN": 1,
    "FEB": 2,
    "MAR": 3,
    "APR": 4,
    "MAY": 5,
    "JUN": 6,
    "JUL": 7,
    "AUG": 8,
    "SEP": 9,
    "OCT": 10,
    "NOV": 11,
    "DEC": 12,
}


def _account_tag(txns: list[Transaction]) -> str:
    if not txns:
        return "UNKNOWN"
    t = txns[0]
    acct = (t.source_account_no or "XXXX")[-4:]
    typ = t.source_account_type or "NA"
    return f"{t.source_bank}-{acct}-{typ}".replace(" ", "")


def _sheet_headers(ws, headers: list[str]) -> None:
    ws.append(headers)


def _autosize(ws) -> None:
    for col in ws.columns:
        max_len = 10
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)) + 2)
        ws.column_dimensions[col[0].column_letter].width = min(max_len, 64)


def _fmt_date(value: date | None) -> str:
    return value.strftime("%d-%b-%Y") if value else ""


def _month_short_from_date(value: date | None) -> str:
    if not value:
        return ""
    month = value.strftime("%b").upper()
    if value.month <= 3:
        return f"{month}({value.strftime('%y')})"
    return month


def _month_start(value: date | None) -> date | None:
    if value is None:
        return None
    return date(value.year, value.month, 1)


def _fiscal_months(all_txns: list[Transaction]) -> list[date]:
    dates = [t.txn_date for t in all_txns if t.txn_date]
    if not dates:
        return []
    first = min(dates)
    last = max(dates)
    months: list[date] = []
    cursor = date(first.year, first.month, 1)
    end = date(last.year, last.month, 1)
    while cursor <= end:
        months.append(cursor)
        if cursor.month == 12:
            cursor = date(cursor.year + 1, 1, 1)
        else:
            cursor = date(cursor.year, cursor.month + 1, 1)
    return months


def _analysis_month_label(value: date) -> str:
    return value.strftime("%b-%y")


def _cons_month_label(value: date) -> str:
    month = value.strftime("%b").upper()
    if value.month <= 3:
        return f"{month} ({value.strftime('%y')})"
    return month


def _month_key_for_pivot(value: date) -> str:
    return value.strftime("%b").upper()


def _extract_case_metadata(
    all_txns: list[Transaction],
    txns_by_account: dict[str, list[Transaction]],
    case_context: dict | None,
    job_name: str | None,
) -> dict[str, str]:
    meta = dict((case_context or {}).get("case_metadata", {}) or {})
    first_txn = all_txns[0] if all_txns else None
    if first_txn:
        meta.setdefault("account_holder", first_txn.source_account_name or job_name or "")
        meta.setdefault("bank_name", first_txn.source_bank or "")
        meta.setdefault("account_number", first_txn.source_account_no or "")
        meta.setdefault("account_type", first_txn.source_account_type or "")
    if len(txns_by_account) >= 2:
        meta.setdefault("bank_name", "MULTIPLE")
        meta.setdefault("account_number", "MULTIPLE")
        meta.setdefault("account_type", "MULTIPLE")
    meta.setdefault("account_holder", job_name or "Borrower")
    meta.setdefault("address", "")
    meta.setdefault("email", "")
    meta.setdefault("pan", "")
    meta.setdefault("mobile", "")
    meta.setdefault("analyst_name", "ANALYST")
    meta.setdefault("other_bank_note", "NO")
    meta.setdefault("feedback", "")
    meta.setdefault("gstr_received", "NO")
    return {str(k): str(v) for k, v in meta.items()}


def _ordered_txns(txns: list[Transaction], reverse: bool = False) -> list[Transaction]:
    return sorted(
        txns,
        key=lambda t: (t.txn_date or date(1900, 1, 1), t.txn_order, t.page_no, t.line_ref),
        reverse=reverse,
    )


def _pivot_type_index(typ: str) -> int:
    try:
        return PIVOT_TYPE_ORDER.index(typ)
    except ValueError:
        return len(PIVOT_TYPE_ORDER)


def _category_fill(typ: str, category: str, category_type_map: dict[str, set[str]]) -> str | None:
    types = category_type_map.get(category, set())
    if "PURCHASE" in types and "SALES" in types:
        return "DDA0DD"
    return TYPE_COLOR.get(typ)


def _set_rupee_columns(ws, cols: tuple[int, ...]) -> None:
    for row in ws.iter_rows(min_row=2):
        for col in cols:
            row[col - 1].number_format = "#,##0"


def _set_lakhs_columns(ws, cols: tuple[int, ...], start_row: int = 2) -> None:
    for row in ws.iter_rows(min_row=start_row):
        for col in cols:
            row[col - 1].number_format = "#,##0.00"


def _first_balance_on_or_after(rows: list[Transaction], month: date, day: int) -> float:
    month_rows = [t for t in _ordered_txns(rows) if t.txn_date and t.txn_date.year == month.year and t.txn_date.month == month.month]
    for txn in month_rows:
        if txn.txn_date and txn.txn_date.day >= day and txn.balance is not None:
            return round(txn.balance, 2)
    return 0.0


def _last_balance_in_month(rows: list[Transaction], month: date) -> float:
    month_rows = [t for t in _ordered_txns(rows) if t.txn_date and t.txn_date.year == month.year and t.txn_date.month == month.month]
    if not month_rows:
        return 0.0
    return round(month_rows[-1].balance or 0.0, 2)


def _account_section_metadata(txns: list[Transaction], fallback: dict[str, str]) -> dict[str, str]:
    meta = dict(fallback)
    if txns:
        first_txn = txns[0]
        meta["account_holder"] = first_txn.source_account_name or meta.get("account_holder", "")
        meta["bank_name"] = first_txn.source_bank or meta.get("bank_name", "")
        meta["account_number"] = first_txn.source_account_no or meta.get("account_number", "")
        meta["account_type"] = first_txn.source_account_type or meta.get("account_type", "")
    return meta


def _write_analysis_section_blueprint(ws, start_row: int, txns: list[Transaction], metadata: dict[str, str], months: list[date]) -> int:
    from openpyxl.styles import Font

    ws.cell(row=start_row, column=1, value="Summary Info")
    ws.cell(row=start_row, column=1).font = Font(bold=True, size=14)

    summary_rows = [
        ("Name of the Account Holder", metadata.get("account_holder", "")),
        ("Address", metadata.get("address", "")),
        ("Email", metadata.get("email", "")),
        ("PAN", metadata.get("pan", "")),
        ("Mobile Number", metadata.get("mobile", "")),
        ("Name of the Bank", metadata.get("bank_name", "")),
        ("Account Number", metadata.get("account_number", "")),
        ("Account Type", metadata.get("account_type", "")),
    ]
    for idx, (label, value) in enumerate(summary_rows, start=start_row + 1):
        ws.cell(row=idx, column=2, value=label)
        ws.cell(row=idx, column=3, value=value)

    month_title_row = start_row + 10
    ws.cell(row=month_title_row, column=1, value="Monthwise Details")
    ws.cell(row=month_title_row, column=1).font = Font(bold=True, size=12)

    month_headers = [_analysis_month_label(m) for m in months]
    for offset, label in enumerate(month_headers, start=3):
        ws.cell(row=month_title_row + 1, column=offset, value=label)
    ws.cell(row=month_title_row + 1, column=3 + len(month_headers), value="TOTAL")

    buckets: dict[date, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    balances: dict[date, list[float]] = defaultdict(list)
    for txn in txns:
        key = _month_start(txn.txn_date)
        if key not in months:
            continue
        if txn.credit > 0:
            buckets[key]["credit_count"] += 1
            buckets[key]["credit_amount"] += txn.credit
        if txn.debit > 0:
            buckets[key]["debit_count"] += 1
            buckets[key]["debit_amount"] += txn.debit
        if txn.classification_primary == "CASH" and txn.credit > 0:
            buckets[key]["cash_credit_count"] += 1
            buckets[key]["cash_credit_amount"] += txn.credit
        if txn.classification_primary == "CASH" and txn.debit > 0:
            buckets[key]["cash_debit_count"] += 1
            buckets[key]["cash_debit_amount"] += txn.debit
        if txn.classification_primary == "RETURN":
            buckets[key]["return_count"] += 1
            narration = txn.cleaned_narration or ""
            if any(token in narration for token in ("INW", "INWARD", "RTGSRET", "I/W")):
                buckets[key]["inward_return_count"] += 1
            if any(token in narration for token in ("ECS", "NACH", "ACH", "OUTWARD", "O/W")):
                buckets[key]["outward_return_count"] += 1
        if txn.balance is not None:
            balances[key].append(txn.balance)

    rows = [
        ("Total No. of Credit Transactions", "credit_count"),
        ("Total Amount of Credit Transactions", "credit_amount"),
        ("Total No. of Debit Transactions", "debit_count"),
        ("Total Amount of Debit Transactions", "debit_amount"),
        ("Total No. of Cash Deposits", "cash_credit_count"),
        ("Total Amount of Cash Deposits", "cash_credit_amount"),
        ("Total No. of Cash Withdrawals", "cash_debit_count"),
        ("Total Amount of Cash Withdrawals", "cash_debit_amount"),
        ("Total No. of Cheque Deposits", "cheque_credit_count"),
        ("Total Amount of Cheque Deposits", "cheque_credit_amount"),
        ("Total No. of Cheque Issues", "cheque_debit_count"),
        ("Total Amount of Cheque Issues", "cheque_debit_amount"),
        ("Total No. of Inward Cheque Bounces", "inward_return_count"),
        ("Total No. of Outward Cheque Bounces", "outward_return_count"),
        ("Penal charges/bounced charges", "return_count"),
        ("Sanction Limit", "sanction_limit"),
        ("Drawing Power Limit", "drawing_power"),
        ("Balance on 10th ", "bal_10"),
        ("Balance on 20th ", "bal_20"),
        ("Balance on last day ", "bal_last"),
        ("Monthwise Peak delay of overdrawing days", "peak_delay"),
        ("Inward Cheque Return (%)", "inward_return_pct"),
        ("Outward Cheque Return (%)", "outward_return_pct"),
        ("Overdrawn Instances No. of days", "overdrawn_days"),
        ("Overdrawn Amount(for all days)", "overdrawn_amount"),
        ("Overdrawn Average Amount", "overdrawn_avg"),
        ("Overdrawn Average as % of Limit", "overdrawn_avg_pct"),
        ("Total Delay in Interest Payment (in days)", "interest_delay"),
        ("Penal Interest > Rs. 500/- No. of times", "penal_interest"),
        ("No of Tranfers from Own Accounts", "own_transfer_count"),
        ("Value of Transfer from Own Accounts (Rs. lakhs)", "own_transfer_value_lakhs"),
        ("Balance on 5th", "bal_5"),
        ("Balance on 15th", "bal_15"),
        ("Balance on 25th", "bal_25"),
    ]
    metric_row_start = month_title_row + 2
    for row_idx, (label, key) in enumerate(rows, start=metric_row_start):
        ws.cell(row=row_idx, column=2, value=label)
        total = 0.0
        for offset, month in enumerate(months, start=3):
            if key == "bal_5":
                value = _first_balance_on_or_after(txns, month, 5)
            elif key == "bal_10":
                value = _first_balance_on_or_after(txns, month, 10)
            elif key == "bal_15":
                value = _first_balance_on_or_after(txns, month, 15)
            elif key == "bal_20":
                value = _first_balance_on_or_after(txns, month, 20)
            elif key == "bal_25":
                value = _first_balance_on_or_after(txns, month, 25)
            elif key == "bal_last":
                value = _last_balance_in_month(txns, month)
            elif key == "own_transfer_count":
                value = sum(1 for t in txns if _month_start(t.txn_date) == month and t.classification_primary == "SIS CON")
            elif key == "own_transfer_value_lakhs":
                amount = sum((t.debit + t.credit) for t in txns if _month_start(t.txn_date) == month and t.classification_primary == "SIS CON")
                value = to_lakhs(amount)
            else:
                value = buckets[month].get(key, 0.0)
            total += value
            ws.cell(row=row_idx, column=offset, value=int(value) if key.endswith("count") else round(value, 2))
        ws.cell(row=row_idx, column=3 + len(month_headers), value=int(total) if key.endswith("count") else round(total, 2))

    balance_rows = [
        ("Min EOD Balance", lambda vals: min(vals) if vals else 0.0),
        ("Max EOD Balance", lambda vals: max(vals) if vals else 0.0),
        ("Average EOD Balance", lambda vals: (sum(vals) / len(vals)) if vals else 0.0),
    ]
    balance_row_start = metric_row_start + len(rows)
    for row_idx, (label, fn) in enumerate(balance_rows, start=balance_row_start):
        ws.cell(row=row_idx, column=2, value=label)
        for offset, month in enumerate(months, start=3):
            ws.cell(row=row_idx, column=offset, value=round(fn(balances.get(month, [])), 2))
        ws.cell(row=row_idx, column=3 + len(months), value=round(fn([v for vals in balances.values() for v in vals]), 2))

    file_header_row = balance_row_start + len(balance_rows) + 3
    file_headers = [
        "File Name",
        "Institution",
        "Account No",
        "Transaction Start Date",
        "Transaction End Date",
        "Name as in Statement",
        "Address as in Statement",
        "Mobile as in Statement",
    ]
    for offset, label in enumerate(file_headers, start=2):
        ws.cell(row=file_header_row, column=offset, value=label)

    dated_rows = [t for t in txns if t.txn_date]
    ws.cell(row=file_header_row + 1, column=2, value=Path(txns[0].source_file).name if txns else "")
    ws.cell(row=file_header_row + 1, column=3, value=metadata.get("bank_name", ""))
    ws.cell(row=file_header_row + 1, column=4, value=metadata.get("account_number", ""))
    ws.cell(row=file_header_row + 1, column=5, value=min((t.txn_date for t in dated_rows), default=None))
    ws.cell(row=file_header_row + 1, column=6, value=max((t.txn_date for t in dated_rows), default=None))
    ws.cell(row=file_header_row + 1, column=7, value=metadata.get("account_holder", ""))
    ws.cell(row=file_header_row + 1, column=8, value=metadata.get("address", ""))
    ws.cell(row=file_header_row + 1, column=9, value=metadata.get("mobile", ""))

    _set_rupee_columns(ws, tuple(range(3, 4 + len(months))))
    return file_header_row + 10


def _create_analysis_sheet_blueprint(
    wb,
    txns_by_account: dict[str, list[Transaction]],
    metadata: dict[str, str],
    months: list[date],
) -> None:
    ws = wb.create_sheet("ANALYSIS")
    next_row = 1
    account_sections = list(txns_by_account.values())
    for idx, txns in enumerate(account_sections):
        next_row = _write_analysis_section_blueprint(ws, next_row, txns, _account_section_metadata(txns, metadata), months)
        if idx < len(account_sections) - 1:
            next_row += 1


def _generic_type_sheet_blueprint(wb, name: str, txns_by_account: dict[str, list[Transaction]]) -> None:
    ws = wb.create_sheet(name)
    wrote_any = False
    for txns in txns_by_account.values():
        rows = [txn for txn in _ordered_txns(txns) if txn.classification_primary == name]
        if not rows:
            continue
        wrote_any = True
        ws.append(["", "", _account_tag(txns), "", "", "", ""])
        _sheet_headers(ws, ["Date", "MONTH", "TYPE", "Cheque_No", "Category", "DR", "CR"])
        for txn in rows:
            ws.append(
                [
                    txn.txn_date,
                    _month_short_from_date(txn.txn_date),
                    txn.classification_primary,
                    txn.cheque_no or "",
                    txn.normalized_party or txn.classification_primary,
                    round(txn.debit, 0),
                    round(txn.credit, 0),
                ]
            )
            ws.cell(row=ws.max_row, column=1).number_format = "DD-MMM-YYYY"
        ws.append(["", "", "", "", "", "", ""])
    if not wrote_any:
        wb.remove(ws)
        return
    _set_rupee_columns(ws, (6, 7))


def _create_analysis_sheet(wb, all_txns: list[Transaction], metadata: dict[str, str], months: list[date]) -> None:
    from openpyxl.styles import Font

    ws = wb.create_sheet("ANALYSIS")
    ws["A1"] = "Summary Info"
    ws["A1"].font = Font(bold=True, size=14)

    summary_rows = [
        ("Name of the Account Holder", metadata.get("account_holder", "")),
        ("Address", metadata.get("address", "")),
        ("Email", metadata.get("email", "")),
        ("PAN", metadata.get("pan", "")),
        ("Mobile Number", metadata.get("mobile", "")),
        ("Name of the Bank", metadata.get("bank_name", "")),
        ("Account Number", metadata.get("account_number", "")),
        ("Account Type", metadata.get("account_type", "")),
    ]
    for idx, (label, value) in enumerate(summary_rows, start=2):
        ws.cell(row=idx, column=2, value=label)
        ws.cell(row=idx, column=3, value=value)

    ws["A11"] = "Monthwise Details"
    ws["A11"].font = Font(bold=True, size=12)

    month_headers = [_analysis_month_label(m) for m in months]
    for offset, label in enumerate(month_headers, start=3):
        ws.cell(row=12, column=offset, value=label)
    ws.cell(row=12, column=3 + len(month_headers), value="TOTAL")

    buckets: dict[date, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    balances: dict[date, list[float]] = defaultdict(list)
    for txn in all_txns:
        key = _month_start(txn.txn_date)
        if key not in months:
            continue
        if txn.credit > 0:
            buckets[key]["credit_count"] += 1
            buckets[key]["credit_amount"] += txn.credit
        if txn.debit > 0:
            buckets[key]["debit_count"] += 1
            buckets[key]["debit_amount"] += txn.debit
        if txn.classification_primary == "CASH" and txn.credit > 0:
            buckets[key]["cash_credit_count"] += 1
            buckets[key]["cash_credit_amount"] += txn.credit
        if txn.classification_primary == "CASH" and txn.debit > 0:
            buckets[key]["cash_debit_count"] += 1
            buckets[key]["cash_debit_amount"] += txn.debit
        if txn.return_flag or txn.classification_primary == "RETURN":
            buckets[key]["return_count"] += 1
        if txn.balance is not None:
            balances[key].append(txn.balance)

    rows = [
        ("Total No. of Credit Transactions", "credit_count"),
        ("Total Amount of Credit Transactions", "credit_amount"),
        ("Total No. of Debit Transactions", "debit_count"),
        ("Total Amount of Debit Transactions", "debit_amount"),
        ("Total No. of Cash Deposits", "cash_credit_count"),
        ("Total Amount of Cash Deposits", "cash_credit_amount"),
        ("Total No. of Cash Withdrawals", "cash_debit_count"),
        ("Total Amount of Cash Withdrawals", "cash_debit_amount"),
        ("Penal charges/bounced charges (count)", "return_count"),
    ]
    for row_idx, (label, key) in enumerate(rows, start=13):
        ws.cell(row=row_idx, column=2, value=label)
        total = 0.0
        for offset, month in enumerate(months, start=3):
            value = buckets[month].get(key, 0.0)
            total += value
            ws.cell(row=row_idx, column=offset, value=int(value) if key.endswith("count") else round(value, 0))
        ws.cell(row=row_idx, column=3 + len(months), value=int(total) if key.endswith("count") else round(total, 0))

    balance_rows = [
        ("Min EOD Balance", lambda vals: min(vals) if vals else 0.0),
        ("Max EOD Balance", lambda vals: max(vals) if vals else 0.0),
        ("Average EOD Balance", lambda vals: (sum(vals) / len(vals)) if vals else 0.0),
    ]
    for row_idx, (label, fn) in enumerate(balance_rows, start=22):
        ws.cell(row=row_idx, column=2, value=label)
        month_values = []
        for offset, month in enumerate(months, start=3):
            value = round(fn(balances.get(month, [])), 0)
            month_values.append(value)
            ws.cell(row=row_idx, column=offset, value=value)
        ws.cell(row=row_idx, column=3 + len(months), value=round(fn([v for vals in balances.values() for v in vals]), 0))

    _set_rupee_columns(ws, tuple(range(3, 4 + len(months))))


def _generic_type_sheet(wb, name: str, rows: list[Transaction]) -> None:
    ws = wb.create_sheet(name)
    _sheet_headers(ws, ["Date", "MONTH", "TYPE", "Cheque_No", "Category", "DR", "CR"])
    for txn in _ordered_txns(rows):
        ws.append(
            [
                _fmt_date(txn.txn_date),
                _month_short_from_date(txn.txn_date),
                txn.classification_primary,
                txn.cheque_no or "",
                txn.normalized_party or txn.classification_primary,
                round(txn.debit, 0),
                round(txn.credit, 0),
            ]
        )
        ws.cell(row=ws.max_row, column=1).number_format = "@"
    _set_rupee_columns(ws, (6, 7))


def _load_lender_rules(case_context: dict | None) -> list[dict]:
    ctx = case_context or {}
    rules_engine = ctx.get("rules_engine", {}) if isinstance(ctx.get("rules_engine", {}), dict) else {}
    entries = []
    raw = ctx.get("lender_grouping")
    if isinstance(raw, list):
        entries.extend(raw)
    raw_nested = rules_engine.get("lender_grouping")
    if isinstance(raw_nested, list):
        entries.extend(raw_nested)
    return [entry for entry in entries if isinstance(entry, dict)]


def _match_lender_rule(txn: Transaction, lender_rules: list[dict]) -> tuple[str | None, bool]:
    narration = (txn.cleaned_narration or txn.raw_narration or "").upper()
    party = (txn.normalized_party or "").upper()
    for entry in lender_rules:
        patterns = entry.get("patterns") or entry.get("match") or entry.get("pattern") or []
        if isinstance(patterns, str):
            patterns = [patterns]
        normalized = [str(p).upper() for p in patterns if str(p).strip()]
        if any(pattern in narration or pattern == party for pattern in normalized):
            return str(entry.get("lender") or entry.get("label") or party).upper(), bool(entry.get("split_by_emi_amount", False))
    return None, False


def _split_bank_fin_loans(txns: list[Transaction]) -> list[tuple[str, list[Transaction]]]:
    by_amount: dict[float, list[Transaction]] = defaultdict(list)
    for txn in _ordered_txns(txns):
        by_amount[round(txn.debit or txn.credit, 2)].append(txn)

    loans: list[tuple[str, list[Transaction]]] = []
    for amount, rows in sorted(by_amount.items(), key=lambda item: item[0], reverse=True):
        by_date: dict[date, list[Transaction]] = defaultdict(list)
        for txn in rows:
            if txn.txn_date:
                by_date[txn.txn_date].append(txn)
        max_per_date = max((len(items) for items in by_date.values()), default=len(rows) or 1)
        buckets: list[list[Transaction]] = [[] for _ in range(max_per_date)]
        for txn_date in sorted(by_date):
            for idx, txn in enumerate(by_date[txn_date]):
                buckets[idx].append(txn)
        for bucket in buckets:
            if bucket:
                loans.append((f"ACC-{len(loans) + 1} (Rs {amount:,.0f} EMI)", bucket))
    return loans


def _create_bank_fin_sheet(wb, rows: list[Transaction], months: list[date], case_context: dict | None) -> None:
    from openpyxl.styles import Font, PatternFill

    ws = wb.create_sheet("BANK FIN")
    _sheet_headers(ws, ["Date", "MONTH", "TYPE", "Cheque_No", "Category", "DR", "CR"])
    lender_rules = _load_lender_rules(case_context)
    grouped: dict[str, list[Transaction]] = defaultdict(list)
    split_flags: dict[str, bool] = {}

    for txn in _ordered_txns(rows):
        lender, split_by_amount = _match_lender_rule(txn, lender_rules)
        key = lender or (txn.normalized_party or txn.classification_primary)
        grouped[key].append(txn)
        split_flags[key] = split_flags.get(key, False) or split_by_amount

    if not grouped:
        return

    for lender in sorted(grouped):
        lender_rows = grouped[lender]
        loan_groups = _split_bank_fin_loans(lender_rows) if split_flags.get(lender, True) else [("ACC-1", lender_rows)]
        for loan_label, loan_rows in loan_groups:
            ws.append(["", "", "", f"{lender} - {loan_label}", "", "", ""])
            heading = ws.max_row
            ws.cell(row=heading, column=4).font = Font(bold=True, color="1F4E78")
            ws.cell(row=heading, column=4).fill = PatternFill(fill_type="solid", fgColor="D9EAF7")

            by_month: dict[date, list[Transaction]] = defaultdict(list)
            for txn in _ordered_txns(loan_rows):
                month = _month_start(txn.txn_date)
                if month:
                    by_month[month].append(txn)
            total_dr = 0.0
            total_cr = 0.0
            for month in months:
                month_rows = by_month.get(month, [])
                if not month_rows:
                    ws.append(["", month.strftime("%b").upper(), "", "", "", "", ""])
                    continue
                for txn in month_rows:
                    total_dr += txn.debit
                    total_cr += txn.credit
                    ws.append(
                        [
                            _fmt_date(txn.txn_date),
                            _month_short_from_date(txn.txn_date),
                            "BANK FIN",
                            txn.cheque_no or "",
                            lender,
                            round(txn.debit, 0),
                            round(txn.credit, 0),
                        ]
                    )
                    ws.cell(row=ws.max_row, column=1).number_format = "@"
            ws.append(["", "", "", "", "TOTAL", round(total_dr, 0), round(total_cr, 0)])
            total_row = ws.max_row
            for col in range(1, 8):
                ws.cell(row=total_row, column=col).font = Font(bold=True)
                ws.cell(row=total_row, column=col).fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
            ws.append(["", "", "", "", "", "", ""])

    _set_rupee_columns(ws, (6, 7))


def _create_pvt_fin_sheet(wb, rows: list[Transaction]) -> None:
    from openpyxl.styles import Font, PatternFill

    ws = wb.create_sheet("PVT FIN")
    _sheet_headers(ws, ["Date", "MONTH", "TYPE", "Cheque_No", "Category", "DR", "CR"])

    if not rows:
        ws.append(["", "", "", "", "NO PRIVATE FINANCE DETECTED", "", ""])
        return

    grouped: dict[str, list[Transaction]] = defaultdict(list)
    for txn in _ordered_txns(rows):
        grouped[txn.normalized_party or "PVT FIN"].append(txn)

    for lender in sorted(grouped):
        lender_rows = grouped[lender]
        credits = [txn for txn in lender_rows if txn.credit > 0]
        debits = [txn for txn in lender_rows if txn.debit > 0]
        given_amount = sum(txn.credit for txn in credits)
        principal_repaid = sum(txn.debit for txn in debits)
        upfront_interest = max(principal_repaid - given_amount, 0.0)
        flat_rate = (upfront_interest / given_amount * 100.0) if given_amount else 0.0

        ws.append(["", "", "", lender, f"GIVEN {given_amount:,.0f} | REPAID {principal_repaid:,.0f} | FLAT {flat_rate:.2f}%", "", ""])
        heading = ws.max_row
        ws.cell(row=heading, column=4).font = Font(bold=True, color="5B2C6F")
        ws.cell(row=heading, column=4).fill = PatternFill(fill_type="solid", fgColor="E8DAEF")

        for txn in _ordered_txns(credits + debits):
            ws.append(
                [
                    _fmt_date(txn.txn_date),
                    _month_short_from_date(txn.txn_date),
                    "PVT FIN",
                    txn.cheque_no or "",
                    lender,
                    round(txn.debit, 0),
                    round(txn.credit, 0),
                ]
            )
            ws.cell(row=ws.max_row, column=1).number_format = "@"
        ws.append(["", "", "", "", "TOTAL", round(principal_repaid, 0), round(given_amount, 0)])
        total_row = ws.max_row
        for col in range(1, 8):
            ws.cell(row=total_row, column=col).font = Font(bold=True)
            ws.cell(row=total_row, column=col).fill = PatternFill(fill_type="solid", fgColor="E8DAEF")
        ws.append(["", "", "", "", "", "", ""])

    _set_rupee_columns(ws, (6, 7))


def _create_pivot_sheet(wb, txns: list[Transaction], tag: str, metadata: dict[str, str]) -> None:
    from openpyxl.styles import Font, PatternFill

    ws = wb.create_sheet(f"PIVOT-{tag}"[:31])
    ws.append([])
    ws.append(["", "", f"{metadata.get('account_holder', '').upper()}\n{tag}"])
    ws.append([])
    ws.append([])
    ws.append([])
    ws.append([f"WORKED BY {metadata.get('analyst_name', 'ANALYST')}"])
    ws.append([])

    months = sorted({_month_start(t.txn_date) for t in txns if t.txn_date})
    month_headers = ["", ""]
    for month in months:
        month_headers.extend([month.strftime("%b").upper(), ""])
    month_headers.extend(["TOTAL", ""])
    ws.append(month_headers)

    headers = ["TYPE", "Category"]
    for month in months:
        headers.extend(["Sum of DR", "Sum of CR"])
    headers.extend(["Sum of DR", "Sum of CR"])
    ws.append(headers)

    pivot: dict[tuple[str, str], dict[date | str, float]] = defaultdict(lambda: defaultdict(float))
    category_type_map: dict[str, set[str]] = defaultdict(set)
    for txn in txns:
        key = (txn.classification_primary, txn.normalized_party or txn.classification_primary)
        month = _month_start(txn.txn_date)
        if month:
            pivot[key][(month, "DR")] += txn.debit
            pivot[key][(month, "CR")] += txn.credit
        pivot[key]["TOTAL_DR"] += txn.debit
        pivot[key]["TOTAL_CR"] += txn.credit
        category_type_map[key[1]].add(txn.classification_primary)

    current_type = None
    totals: dict[date | str, float] | None = None
    for typ, category in sorted(pivot.keys(), key=lambda item: (_pivot_type_index(item[0]), item[1])):
        if current_type is not None and typ != current_type and totals is not None:
            row = [f"{current_type} Total", ""]
            for month in months:
                row.extend([to_lakhs(totals.get((month, "DR"), 0.0)), to_lakhs(totals.get((month, "CR"), 0.0))])
            row.extend([to_lakhs(totals.get("TOTAL_DR", 0.0)), to_lakhs(totals.get("TOTAL_CR", 0.0))])
            ws.append(row)
            for col in range(1, len(row) + 1):
                ws.cell(row=ws.max_row, column=col).font = Font(bold=True)
        if typ != current_type:
            current_type = typ
            totals = defaultdict(float)
        row = [typ, category]
        vals = pivot[(typ, category)]
        for month in months:
            dr = vals.get((month, "DR"), 0.0)
            cr = vals.get((month, "CR"), 0.0)
            row.extend([to_lakhs(dr), to_lakhs(cr)])
            totals[(month, "DR")] += dr
            totals[(month, "CR")] += cr
        row.extend([to_lakhs(vals.get("TOTAL_DR", 0.0)), to_lakhs(vals.get("TOTAL_CR", 0.0))])
        totals["TOTAL_DR"] += vals.get("TOTAL_DR", 0.0)
        totals["TOTAL_CR"] += vals.get("TOTAL_CR", 0.0)
        ws.append(row)
        fill = _category_fill(typ, category, category_type_map)
        if fill:
            ws.cell(row=ws.max_row, column=2).fill = PatternFill(fill_type="solid", fgColor=fill)

    if current_type is not None and totals is not None:
        row = [f"{current_type} Total", ""]
        for month in months:
            row.extend([to_lakhs(totals.get((month, "DR"), 0.0)), to_lakhs(totals.get((month, "CR"), 0.0))])
        row.extend([to_lakhs(totals.get("TOTAL_DR", 0.0)), to_lakhs(totals.get("TOTAL_CR", 0.0))])
        ws.append(row)
        for col in range(1, len(row) + 1):
            ws.cell(row=ws.max_row, column=col).font = Font(bold=True)

    _set_lakhs_columns(ws, tuple(range(3, len(headers) + 1)), start_row=10)


def _create_xns_sheet(wb, txns: list[Transaction], tag: str) -> None:
    ws = wb.create_sheet(f"XNS-{tag}"[:31])
    _sheet_headers(ws, ["Sl No", "Date", "MONTH", "TYPE", "Cheque_No", "Category", "Description", "DR", "CR", "Balance"])
    for idx, txn in enumerate(_ordered_txns(txns, reverse=True), start=1):
        ws.append(
            [
                idx,
                _fmt_date(txn.txn_date),
                _month_short_from_date(txn.txn_date),
                txn.classification_primary,
                txn.cheque_no or "",
                txn.normalized_party or txn.classification_primary,
                txn.raw_narration,
                to_lakhs(txn.debit),
                to_lakhs(txn.credit),
                round(txn.balance, 0) if txn.balance is not None else "",
            ]
        )
        ws.cell(row=ws.max_row, column=2).number_format = "@"
        ws.cell(row=ws.max_row, column=10).number_format = "#,##0"
    _set_lakhs_columns(ws, (8, 9))
    ws.auto_filter.ref = ws.dimensions


def _create_cons_sheet(wb, txns_by_account: dict[str, list[Transaction]], months: list[date]) -> None:
    ws = wb.create_sheet("CONS")
    ws.append([])
    ws.append([])
    ws.append([])
    ws.append([])
    headers = ["", "", "", "MONTH"]
    account_keys = sorted(txns_by_account.keys())
    for key in account_keys:
        headers.extend([f"{key}\nPURCHASE", f"{key}\nSALES"])
    headers.extend(["PURCHASE\nTOTAL", "SALES\nTOTAL"])
    ws.append(headers)
    ws.append([])
    for month in months:
        row = ["", "", "", month.strftime("%b").upper()]
        total_purchase = 0.0
        total_sales = 0.0
        for key in account_keys:
            rows = txns_by_account[key]
            purchase = sum(t.debit for t in rows if _month_start(t.txn_date) == month and t.classification_primary == "PURCHASE")
            sales = sum(t.credit for t in rows if _month_start(t.txn_date) == month and t.classification_primary == "SALES")
            total_purchase += purchase
            total_sales += sales
            row.extend([round(purchase, 0), round(sales, 0)])
        row.extend([round(total_purchase, 0), round(total_sales, 0)])
        ws.append(row)
    _set_rupee_columns(ws, tuple(range(5, 7 + 2 * len(account_keys))))


def _turnover_trend(all_txns: list[Transaction], months: list[date]) -> str:
    sales_by_month = {month: 0.0 for month in months}
    for txn in all_txns:
        month = _month_start(txn.txn_date)
        if month in sales_by_month and txn.classification_primary == "SALES":
            sales_by_month[month] += txn.credit
    nonzero = [value for value in sales_by_month.values() if value > 0]
    if len(nonzero) < 2:
        return "FLUCTUATING"
    first = next(value for value in sales_by_month.values() if value > 0)
    last = next(value for value in reversed(list(sales_by_month.values())) if value > 0)
    if last > first * 1.05:
        return "INCREASE"
    if last < first * 0.95:
        return "DECREASE"
    return "FLUCTUATING"


def _bank_fin_missing_months(rows: list[Transaction], months: list[date]) -> bool:
    lender_rows: dict[tuple[str, float], set[date]] = defaultdict(set)
    for txn in rows:
        if txn.debit <= 0 or not txn.txn_date:
            continue
        lender_rows[(txn.normalized_party or "BANK FIN", round(txn.debit, 2))].add(_month_start(txn.txn_date))
    for seen in lender_rows.values():
        active = [month for month in months if month in seen]
        if len(active) >= 2:
            start = months.index(active[0])
            end = months.index(active[-1])
            if any(month not in seen for month in months[start : end + 1]):
                return True
    return False


def _create_final_sheet(wb, all_txns: list[Transaction], metadata: dict[str, str], months: list[date]) -> None:
    from openpyxl.styles import Font

    ws = wb.create_sheet("FINAL")
    ws["G2"] = "Checklist"
    ws["G2"].font = Font(bold=True, size=13)

    total_sales = sum(t.credit for t in all_txns if t.classification_primary == "SALES")
    total_bank_fin = sum(t.debit for t in all_txns if t.classification_primary == "BANK FIN")
    total_pvt_fin = sum(t.debit for t in all_txns if t.classification_primary == "PVT FIN")
    inward_return_count = sum(
        1
        for t in all_txns
        if t.classification_primary == "RETURN" and any(token in (t.cleaned_narration or "") for token in ("INW", "INWARD", "CHQ RET"))
    )
    ecs_return = any(
        t.classification_primary == "RETURN" and any(token in (t.cleaned_narration or "") for token in ("ECS", "NACH", "ACH"))
        for t in all_txns
    )
    irregular_finance = _bank_fin_missing_months(
        [t for t in all_txns if t.classification_primary == "BANK FIN"],
        months,
    )
    rows = [
        ("INWARD CHEQUE RETURN", f"YES - {inward_return_count} instances" if inward_return_count else "NO"),
        ("ECS CHEQUE RETURN", "YES" if ecs_return else "NO"),
        ("IRREGULAR REPAYMENT FINANCE", "YES" if irregular_finance else "NO"),
        ("PRIVATE FINANCE ON DATE", "YES" if total_pvt_fin and not irregular_finance else "NO"),
        ("PRIVATE FINANCE 10%", "YES ABOVE 10%" if total_sales and total_pvt_fin > total_sales * 0.1 else "NO ABOVE 10%"),
        ("BANK FINANCE 10%", "YES ABOVE 10%" if total_sales and total_bank_fin > total_sales * 0.1 else "NO ABOVE 10%"),
        ("GSTR RECEIVED", metadata.get("gstr_received", "NO").upper()),
        ("TURN OVER (MONTHLY) INCREASE / DECREASE", _turnover_trend(all_txns, months)),
        ("Is there any other bank statement?", metadata.get("other_bank_note", "NO")),
        ("OUR FEED BACK", metadata.get("feedback", "")),
    ]
    for idx, (label, value) in enumerate(rows, start=4):
        ws.cell(row=idx, column=7, value=label)
        ws.cell(row=idx, column=8, value=value)
        ws.cell(row=idx, column=7).font = Font(bold=True)


def _create_bank_fin_sheet_blueprint(wb, txns_by_account: dict[str, list[Transaction]], months: list[date], case_context: dict | None) -> None:
    from openpyxl.styles import Font, PatternFill

    ws = wb.create_sheet("BANK FIN")
    lender_rules = _load_lender_rules(case_context)
    wrote_any = False
    for txns in txns_by_account.values():
        rows = [txn for txn in _ordered_txns(txns) if txn.classification_primary == "BANK FIN"]
        if not rows:
            continue
        wrote_any = True
        ws.append(["", "", _account_tag(txns), "", "", "", ""])
        _sheet_headers(ws, ["Date", "MONTH", "TYPE", "Cheque_No", "Category", "DR", "CR"])

        grouped: dict[str, list[Transaction]] = defaultdict(list)
        split_flags: dict[str, bool] = {}
        for txn in rows:
            lender, split_by_amount = _match_lender_rule(txn, lender_rules)
            key = lender or (txn.normalized_party or txn.classification_primary)
            grouped[key].append(txn)
            split_flags[key] = split_flags.get(key, False) or split_by_amount

        for lender in sorted(grouped):
            lender_rows = grouped[lender]
            loan_groups = _split_bank_fin_loans(lender_rows) if split_flags.get(lender, True) else [("ACC-1", lender_rows)]
            for loan_label, loan_rows in loan_groups:
                ws.append(["", "", "", loan_label.replace("-", " "), "", "", ""])
                heading = ws.max_row
                ws.cell(row=heading, column=4).font = Font(bold=True, color="1F4E78")
                ws.cell(row=heading, column=4).fill = PatternFill(fill_type="solid", fgColor="D9EAF7")

                by_month: dict[date, list[Transaction]] = defaultdict(list)
                for txn in _ordered_txns(loan_rows):
                    month = _month_start(txn.txn_date)
                    if month:
                        by_month[month].append(txn)
                total_dr = 0.0
                total_cr = 0.0
                for month in months:
                    month_rows = by_month.get(month, [])
                    if not month_rows:
                        ws.append(["", _month_short_from_date(month), "", "", "", "", ""])
                        continue
                    for txn in month_rows:
                        total_dr += txn.debit
                        total_cr += txn.credit
                        ws.append(
                            [
                                txn.txn_date,
                                _month_short_from_date(txn.txn_date),
                                "BANK FIN",
                                txn.cheque_no or "",
                                lender,
                                round(txn.debit, 0),
                                round(txn.credit, 0),
                            ]
                        )
                        ws.cell(row=ws.max_row, column=1).number_format = "DD-MMM-YYYY"
                ws.append(["", "", "", "", "TOTAL", round(total_dr, 0), round(total_cr, 0)])
                total_row = ws.max_row
                for col in range(1, 8):
                    ws.cell(row=total_row, column=col).font = Font(bold=True)
                    ws.cell(row=total_row, column=col).fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
                ws.append(["", "", "", "", "", "", ""])
    if not wrote_any:
        wb.remove(ws)
        return
    _set_rupee_columns(ws, (6, 7))


def _create_pvt_fin_sheet_blueprint(wb, txns_by_account: dict[str, list[Transaction]]) -> None:
    from openpyxl.styles import Font, PatternFill

    ws = wb.create_sheet("PVT FIN")
    wrote_any = False
    for txns in txns_by_account.values():
        rows = [txn for txn in _ordered_txns(txns) if txn.classification_primary == "PVT FIN"]
        if not rows:
            continue
        wrote_any = True
        ws.append(["", "", _account_tag(txns), "", "", "", ""])
        _sheet_headers(ws, ["Date", "MONTH", "TYPE", "Cheque_No", "Category", "DR", "CR"])

        grouped: dict[str, list[Transaction]] = defaultdict(list)
        for txn in rows:
            grouped[txn.normalized_party or "PVT FIN"].append(txn)

        for idx, lender in enumerate(sorted(grouped), start=1):
            lender_rows = grouped[lender]
            credits = [txn for txn in lender_rows if txn.credit > 0]
            debits = [txn for txn in lender_rows if txn.debit > 0]
            given_amount = sum(txn.credit for txn in credits)
            principal_repaid = sum(txn.debit for txn in debits)

            ws.append(["", "", "", f"ACC {idx}", "", "", ""])
            heading = ws.max_row
            ws.cell(row=heading, column=4).font = Font(bold=True, color="5B2C6F")
            ws.cell(row=heading, column=4).fill = PatternFill(fill_type="solid", fgColor="E8DAEF")

            for txn in _ordered_txns(credits + debits):
                ws.append(
                    [
                        txn.txn_date,
                        _month_short_from_date(txn.txn_date),
                        "PVT FIN",
                        txn.cheque_no or "",
                        lender,
                        round(txn.debit, 0),
                        round(txn.credit, 0),
                    ]
                )
                ws.cell(row=ws.max_row, column=1).number_format = "DD-MMM-YYYY"
            ws.append(["", "", "", "", "TOTAL", round(principal_repaid, 0), round(given_amount, 0)])
            total_row = ws.max_row
            for col in range(1, 8):
                ws.cell(row=total_row, column=col).font = Font(bold=True)
                ws.cell(row=total_row, column=col).fill = PatternFill(fill_type="solid", fgColor="E8DAEF")
            ws.append(["", "", "", "", "", "", ""])

    if not wrote_any:
        ws.append(["", "", "", "", "NO PRIVATE FINANCE DETECTED", "", ""])
        return
    _set_rupee_columns(ws, (6, 7))


def _create_pivot_sheet_blueprint(wb, txns: list[Transaction], tag: str, metadata: dict[str, str]) -> None:
    from openpyxl.styles import Font, PatternFill

    ws = wb.create_sheet(f"PIVOT-{tag}"[:31])
    ws["D1"] = f"{metadata.get('account_holder', '').upper()}\n{tag}"
    ws["A5"] = f"ANALYSED BY : {metadata.get('analyst_name', 'ANALYST')}"
    ws["C6"] = "MONTH"
    ws["D6"] = "Values"

    months = sorted({_month_start(t.txn_date) for t in txns if t.txn_date})
    month_headers = ["", ""]
    for month in months:
        month_headers.extend([month.strftime("%b").upper(), ""])
    month_headers.extend(["TOTAL", ""])
    for col, value in enumerate(month_headers, start=1):
        ws.cell(row=7, column=col, value=value)

    headers = ["TYPE", "Category"]
    for month in months:
        headers.extend(["Sum of DR", "Sum of CR"])
    headers.extend(["Sum of DR", "Sum of CR"])
    for col, value in enumerate(headers, start=1):
        ws.cell(row=8, column=col, value=value)

    pivot: dict[tuple[str, str], dict[date | str, float]] = defaultdict(lambda: defaultdict(float))
    category_type_map: dict[str, set[str]] = defaultdict(set)
    for txn in txns:
        key = (txn.classification_primary, txn.normalized_party or txn.classification_primary)
        month = _month_start(txn.txn_date)
        if month:
            pivot[key][(month, "DR")] += txn.debit
            pivot[key][(month, "CR")] += txn.credit
        pivot[key]["TOTAL_DR"] += txn.debit
        pivot[key]["TOTAL_CR"] += txn.credit
        category_type_map[key[1]].add(txn.classification_primary)

    current_type = None
    totals: dict[date | str, float] | None = None
    for typ, category in sorted(pivot.keys(), key=lambda item: (_pivot_type_index(item[0]), item[1])):
        if current_type is not None and typ != current_type and totals is not None:
            row = [f"{current_type} Total", ""]
            for month in months:
                row.extend([to_lakhs(totals.get((month, "DR"), 0.0)), to_lakhs(totals.get((month, "CR"), 0.0))])
            row.extend([to_lakhs(totals.get("TOTAL_DR", 0.0)), to_lakhs(totals.get("TOTAL_CR", 0.0))])
            next_row = ws.max_row + 1
            for col, value in enumerate(row, start=1):
                ws.cell(row=next_row, column=col, value=value)
                ws.cell(row=next_row, column=col).font = Font(bold=True)
        if typ != current_type:
            current_type = typ
            totals = defaultdict(float)
        row = [typ, category]
        vals = pivot[(typ, category)]
        for month in months:
            dr = vals.get((month, "DR"), 0.0)
            cr = vals.get((month, "CR"), 0.0)
            row.extend([to_lakhs(dr), to_lakhs(cr)])
            totals[(month, "DR")] += dr
            totals[(month, "CR")] += cr
        row.extend([to_lakhs(vals.get("TOTAL_DR", 0.0)), to_lakhs(vals.get("TOTAL_CR", 0.0))])
        totals["TOTAL_DR"] += vals.get("TOTAL_DR", 0.0)
        totals["TOTAL_CR"] += vals.get("TOTAL_CR", 0.0)
        next_row = ws.max_row + 1
        for col, value in enumerate(row, start=1):
            ws.cell(row=next_row, column=col, value=value)
        fill = _category_fill(typ, category, category_type_map)
        if fill:
            ws.cell(row=next_row, column=2).fill = PatternFill(fill_type="solid", fgColor=fill)

    if current_type is not None and totals is not None:
        row = [f"{current_type} Total", ""]
        for month in months:
            row.extend([to_lakhs(totals.get((month, "DR"), 0.0)), to_lakhs(totals.get((month, "CR"), 0.0))])
        row.extend([to_lakhs(totals.get("TOTAL_DR", 0.0)), to_lakhs(totals.get("TOTAL_CR", 0.0))])
        next_row = ws.max_row + 1
        for col, value in enumerate(row, start=1):
            ws.cell(row=next_row, column=col, value=value)
            ws.cell(row=next_row, column=col).font = Font(bold=True)

    _set_lakhs_columns(ws, tuple(range(3, len(headers) + 1)), start_row=9)


def _create_xns_sheet_blueprint(wb, txns: list[Transaction], tag: str) -> None:
    ws = wb.create_sheet(f"XNS-{tag}"[:31])
    _sheet_headers(ws, ["Sl. No.", "Date", "MONTH", "TYPE", "Cheque_No", "Category", "Description", "DR", "CR", "Balance"])
    for idx, txn in enumerate(_ordered_txns(txns), start=1):
        ws.append(
            [
                idx,
                txn.txn_date,
                txn.txn_date.strftime("%b").upper() if txn.txn_date else "",
                txn.classification_primary,
                txn.cheque_no or "",
                txn.normalized_party or txn.classification_primary,
                txn.raw_narration,
                to_lakhs(txn.debit),
                to_lakhs(txn.credit),
                round(txn.balance, 0) if txn.balance is not None else "",
            ]
        )
        ws.cell(row=ws.max_row, column=2).number_format = "DD-MMM-YYYY"
        ws.cell(row=ws.max_row, column=10).number_format = "#,##0"
    _set_lakhs_columns(ws, (8, 9))
    ws.auto_filter.ref = ws.dimensions


def _create_cons_sheet_blueprint(wb, txns_by_account: dict[str, list[Transaction]], months: list[date]) -> None:
    ws = wb.create_sheet("CONS")
    ws.append([])
    account_keys = list(txns_by_account.keys())
    account_names = [txns_by_account[key][0].source_account_name or key for key in account_keys if txns_by_account.get(key)]
    title = f"{' & '.join(account_names)} - {len(account_keys)} ACC PURCHASE &SALES" if account_names else ""
    ws.append(["", "", "", title])
    ws.append([])
    ws.append([])
    headers = ["", "", "", "MONTH"]
    for key in account_keys:
        name = txns_by_account[key][0].source_account_name or key
        headers.append(f"{name}\n{_account_tag(txns_by_account[key])}\nPURCHASE")
    headers.append("PURCHASE\n TOTAL")
    for key in account_keys:
        name = txns_by_account[key][0].source_account_name or key
        headers.append(f"{name}\n{_account_tag(txns_by_account[key])}\nSALES")
    headers.append("SALES \nTOTAL")
    ws.append(headers)
    ws.append([])
    for month in months:
        row = ["", "", "", _cons_month_label(month)]
        total_purchase = 0.0
        total_sales = 0.0
        for key in account_keys:
            rows = txns_by_account[key]
            purchase = sum(t.debit for t in rows if _month_start(t.txn_date) == month and t.classification_primary == "PURCHASE")
            total_purchase += purchase
            row.append(to_lakhs(purchase))
        row.append(to_lakhs(total_purchase))
        for key in account_keys:
            rows = txns_by_account[key]
            sales = sum(t.credit for t in rows if _month_start(t.txn_date) == month and t.classification_primary == "SALES")
            total_sales += sales
            row.append(to_lakhs(sales))
        row.append(to_lakhs(total_sales))
        ws.append(row)
        ws.append([""] * len(headers))
    _set_lakhs_columns(ws, tuple(range(5, 7 + (2 * len(account_keys)))))


def _create_final_sheet_blueprint(wb, all_txns: list[Transaction], metadata: dict[str, str], months: list[date]) -> None:
    ws = wb.create_sheet("FINAL")
    total_sales = sum(t.credit for t in all_txns if t.classification_primary == "SALES")
    total_bank_fin = sum(t.debit for t in all_txns if t.classification_primary == "BANK FIN")
    total_pvt_fin = sum(t.debit for t in all_txns if t.classification_primary == "PVT FIN")
    inward_return_count = sum(
        1
        for t in all_txns
        if t.classification_primary == "RETURN" and any(token in (t.cleaned_narration or "") for token in ("INW", "INWARD", "CHQ RET"))
    )
    ecs_return = any(
        t.classification_primary == "RETURN" and any(token in (t.cleaned_narration or "") for token in ("ECS", "NACH", "ACH"))
        for t in all_txns
    )
    irregular_finance = _bank_fin_missing_months([t for t in all_txns if t.classification_primary == "BANK FIN"], months)
    rows = [
        ("INWARD CHEQUE RETURN", "YES" if inward_return_count else "NO"),
        ("ECS CHEQUE RETURN", "YES" if ecs_return else "NO"),
        ("IRREUGULAR REPAYMENT FINANCE", "YES" if irregular_finance else "NO"),
        ("PRIVATE FINANCE ON DATE", "YES" if total_pvt_fin and not irregular_finance else "NO"),
        ("PRIVATE FINANCE  10%", "YES" if total_sales and total_pvt_fin > total_sales * 0.1 else "NO"),
        ("BANK FINANCE 10%", "YES" if total_sales and total_bank_fin > total_sales * 0.1 else "NO"),
        ("GSTR RECEIVED", metadata.get("gstr_received", "NO").upper()),
        ("TURN OVER (MONTHLY) INCREASE / DECREASE", _turnover_trend(all_txns, months)),
        ("Is there any other bank statement?", metadata.get("other_bank_note", "NO")),
        ("OUR FEED BACK", metadata.get("feedback", "")),
    ]
    for idx, (label, value) in enumerate(rows, start=14):
        ws.cell(row=idx, column=7, value=label)
        ws.cell(row=idx, column=8, value=value)


def write_workbook(
    output_path: Path,
    txns_by_account: dict[str, list[Transaction]],
    *,
    job_name: str | None = None,
    case_context: dict | None = None,
) -> Path:
    try:
        from openpyxl import Workbook
    except Exception as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required for workbook export") from exc

    wb = Workbook()
    wb.remove(wb.active)

    all_txns = [txn for rows in txns_by_account.values() for txn in rows]
    months = _fiscal_months(all_txns)
    metadata = _extract_case_metadata(all_txns, txns_by_account, case_context, job_name)

    _create_analysis_sheet_blueprint(wb, txns_by_account, metadata, months)
    _generic_type_sheet_blueprint(wb, "ODD FIG", txns_by_account)
    _generic_type_sheet_blueprint(wb, "DOUBT", txns_by_account)
    _generic_type_sheet_blueprint(wb, "NAMES", txns_by_account)
    _create_bank_fin_sheet_blueprint(wb, txns_by_account, months, case_context)
    _create_pvt_fin_sheet_blueprint(wb, txns_by_account)
    _generic_type_sheet_blueprint(wb, "RETURN", txns_by_account)

    for account_key, txns in txns_by_account.items():
        tag = _account_tag(txns)
        account_metadata = _account_section_metadata(txns, metadata)
        _create_pivot_sheet_blueprint(wb, txns, tag, account_metadata)
        _create_xns_sheet_blueprint(wb, txns, tag)

    if len(txns_by_account) >= 2:
        _create_cons_sheet_blueprint(wb, txns_by_account, months)

    _create_final_sheet_blueprint(wb, all_txns, metadata, months)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        _autosize(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
